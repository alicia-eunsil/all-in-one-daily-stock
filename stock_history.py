"""
File: stock_history.py
Version: v2.1.0
Role: 한국투자증권 API를 호출해 주식·ETF 히스토리를 가져오고 엑셀의 시세/지수 시트를 갱신한다.
# 메모: v2.1.0 - KR_Stocks_Individual에 개인/외국인/기관계 순매수(일별) 시트 업데이트 추가
"""

import json
import requests
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import time

NETBUY_SHEET_BY_KEY = {
    "personal": "순매수_개인",
    "foreign": "순매수_외국인",
    "institution": "순매수_기관계",
}


# =========================
# 0. 설정/공통 유틸 함수들
# =========================

def load_api_secrets(file_path='secrets.json'):
    """API 키와 시크릿을 파일에서 로드"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"에러: {file_path} 파일을 찾을 수 없습니다.")
        return None


def load_file_config(file_path='stock_file_map.json'):
    """자산별 엑셀 파일 매핑 정보 로드"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"에러: {file_path} 파일을 찾을 수 없습니다.")
        return None


def get_token(api_key, api_secret, domain):
    """한국투자증권 API 토큰 발급 요청"""
    url = f"{domain}/oauth2/tokenP"

    headers = {
        "content-type": "application/json",
        "appKey": api_key,
        "appSecret": api_secret
    }

    data = {
        "grant_type": "client_credentials",
        "appkey": api_key,
        "appsecret": api_secret
    }

    try:
        resp = requests.post(url, headers=headers, json=data)

        if resp.status_code != 200:
            print(f"❌ 토큰 요청 실패: HTTP {resp.status_code}")
            print(resp.text)
            return None

        token_data = resp.json()
        if not token_data or 'access_token' not in token_data:
            print("❌ 토큰 정보가 응답에 없습니다")
            return None

        print("✅ 토큰 발급 성공!")
        return token_data

    except requests.exceptions.RequestException as e:
        print(f"❌ 토큰 요청 실패: {str(e)}")
        if hasattr(e, 'response') and e.response is not None:
            print(f"서버 응답: {e.response.text}")
        return None


def _to_int_safe(v):
    if v is None:
        return None
    s = str(v).replace(",", "").strip()
    if s == "":
        return None
    try:
        return int(float(s))
    except Exception:
        return None


# =========================
# 1. 시세 조회 함수들
# =========================

def fetch_stock_daily_history(access_token, domain, symbol, start_date, end_date,
                              app_key=None, app_secret=None):
    """
    국내 주식/ETF 기간별 시세 (일별)
    /uapi/domestic-stock/v1/quotations/inquire-daily-itemchartprice
    """
    endpoint = f"{domain}/uapi/domestic-stock/v1/quotations/inquire-daily-itemchartprice"

    params = {
        "FID_COND_MRKT_DIV_CODE": "J",    # 주식 시장 구분
        "FID_INPUT_ISCD": symbol,         # 종목코드
        "FID_PERIOD_DIV_CODE": "D",       # 기간 구분 (일)
        "FID_ORG_ADJ_PRC": "1",           # 수정주가 여부
        "FID_INPUT_DATE_1": start_date,   # 조회 시작일
        "FID_INPUT_DATE_2": end_date,     # 조회 종료일
        "FID_COMP_ICD": symbol,
    }

    headers = {
        "content-type": "application/json; charset=utf-8",
        "authorization": f"Bearer {access_token}",
        "appkey": app_key,
        "appsecret": app_secret,
        "tr_id": "FHKST03010100",     # 주식 일별 시세
        "custtype": "P",
        "seq_no": "0",
        "locale": "ko_KR",
    }

    try:
        resp = requests.get(endpoint, headers=headers, params=params, timeout=10)

        if resp.status_code != 200:
            print(f"❌ 국내 시세 HTTP {resp.status_code} 에러: {resp.text}")
            return None

        data = resp.json()
        if not data or 'output2' not in data or not data['output2']:
            # print("❌ 국내 시세 데이터가 비어있습니다")
            return None

        daily_data = []
        for item in data['output2']:
            daily_data.append({
                'date': item.get('stck_bsop_date', ''),
                'open': int(item.get('stck_oprc', '0') or 0),
                'high': int(item.get('stck_hgpr', '0') or 0),
                'low': int(item.get('stck_lwpr', '0') or 0),
                'close': int(item.get('stck_clpr', '0') or 0),
                'volume': int(item.get('acml_vol', '0') or 0)
            })

        # 과거 → 최신 정렬
        daily_data.sort(key=lambda x: x['date'])
        return daily_data

    except Exception as e:
        print(f"❌ 국내 시세 조회 중 에러: {str(e)}")
        return None


def fetch_overseas_daily_history(access_token, domain, market_code, symbol,
                                 start_date, end_date, app_key=None, app_secret=None):
    """
    해외 주식/ETF 기간별 시세 (일/주/월)
    /uapi/overseas-price/v1/quotations/dailyprice
    - 여기서는 일봉(GUBN=0), BYMD=end_date 기준으로 최근 100개 받아서
      date 필터링(start_date~end_date)만 적용
    """
    endpoint = f"{domain}/uapi/overseas-price/v1/quotations/dailyprice"

    headers = {
        "content-type": "application/json; charset=utf-8",
        "authorization": f"Bearer {access_token}",
        "appkey": app_key,
        "appsecret": app_secret,
        "tr_id": "HHDFS76240000",   # 해외 주식 기간별 시세
        "custtype": "P",
    }

    params = {
        "AUTH": "",
        "EXCD": market_code,   # 예: "NAS"
        "SYMB": symbol,        # 예: "AAPL"
        "GUBN": "0",           # 0: 일, 1: 주, 2: 월
        "BYMD": end_date,      # 기준일 (이 날 포함 과거 방향 최대 100개)
        "MODP": "0",           # 0: 원주가, 1: 수정주가
        # "KEYB": ""           # 연속조회시 사용, 여기서는 생략
    }

    try:
        resp = requests.get(endpoint, headers=headers, params=params, timeout=10)

        if resp.status_code != 200:
            print(f"❌ 해외 시세 HTTP {resp.status_code} 에러: {resp.text}")
            return None

        data = resp.json()
        rows = data.get("output2")
        if not rows:
            # print("❌ 해외 시세 데이터가 비어있습니다")
            return None

        daily_data = []
        for item in rows:
            d = item.get("xymd")   # 날짜 (YYYYMMDD)
            if not d:
                continue

            # 필드명은 실제 응답에 따라 필요시 한번 확인
            daily_data.append({
                "date": d,
                "open": float(item.get("open", 0) or 0),
                "high": float(item.get("high", 0) or 0),
                "low": float(item.get("low", 0) or 0),
                "close": float(item.get("clos", 0) or 0),
                "volume": int(item.get("tvol", 0) or 0),
            })

        # 과거 → 최신
        daily_data.sort(key=lambda x: x["date"])

        # start_date~end_date로 필터링
        if start_date:
            daily_data = [d for d in daily_data if start_date <= d["date"] <= end_date]

        return daily_data

    except Exception as e:
        print(f"❌ 해외 시세 조회 중 에러: {str(e)}")
        return None


def fetch_investor_netbuy_history(access_token, domain, symbol, start_date,
                                  app_key=None, app_secret=None):
    """
    국내 종목별 투자자매매동향(일별) 조회
    /uapi/domestic-stock/v1/quotations/investor-trade-by-stock-daily
    - 개인/외국인/기관계 순매수 수량만 사용
    """
    endpoint = f"{domain}/uapi/domestic-stock/v1/quotations/investor-trade-by-stock-daily"

    params = {
        "FID_COND_MRKT_DIV_CODE": "J",
        "FID_INPUT_ISCD": symbol,
        "FID_INPUT_DATE_1": start_date,
        "FID_ORG_ADJ_PRC": "",
        "FID_ETC_CLS_CODE": "",
    }

    headers = {
        "content-type": "application/json; charset=utf-8",
        "authorization": f"Bearer {access_token}",
        "appkey": app_key,
        "appsecret": app_secret,
        "tr_id": "FHPTJ04160001",
        "custtype": "P",
    }

    try:
        resp = requests.get(endpoint, headers=headers, params=params, timeout=10)

        if resp.status_code != 200:
            print(f"❌ 순매수 HTTP {resp.status_code} 에러 ({symbol}): {resp.text}")
            return None

        data = resp.json()
        rows = data.get("output2")
        if not rows:
            return None

        history = []
        for row in rows:
            d = row.get("stck_bsop_date", "")
            if not d:
                continue
            history.append({
                "date": d,
                "personal": _to_int_safe(row.get("prsn_ntby_qty")),
                "foreign": _to_int_safe(row.get("frgn_ntby_qty")),
                "institution": _to_int_safe(row.get("orgn_ntby_qty")),
            })

        history.sort(key=lambda x: x["date"])
        if start_date:
            history = [h for h in history if h["date"] >= start_date]

        return history or None
    except Exception as e:
        print(f"❌ 순매수 조회 중 에러 ({symbol}): {str(e)}")
        return None


# =========================
# 2. 엑셀 관련 함수들
# =========================

def load_stock_list(filename, market="KR"):
    """
    Excel 파일에서 종목 목록을 읽어옵니다.
    - market="KR" → 코드 6자리 zfill
    - market="US" → 코드 그대로 사용
    """
    try:
        wb = openpyxl.load_workbook(filename)
        if "종목" not in wb.sheetnames:
            print(f"\n❌ Excel 파일({filename})에 '종목' 시트가 없습니다.")
            return None
        sheet = wb["종목"]

        stocks = []
        for row in sheet.iter_rows(min_row=2):  # 헤더 제외
            if row[0].value and row[1].value:
                raw_code = str(row[1].value).strip()
                if market == "KR":
                    code = raw_code.zfill(6)
                else:
                    code = raw_code

                stocks.append({
                    'name': row[0].value,
                    'code': code
                })

        print(f"\n[{filename}]에서 읽어온 종목 목록 ({market}):")
        for stock in stocks:
            print(f"  • {stock['name']} (코드: {stock['code']})")

        return stocks

    except Exception as e:
        print(f"\n❌ Excel 파일 읽기 실패({filename}): {str(e)}")
        return None


def save_history_to_excel(data_list, filename, market="KR"):
    """
    각 종목의 일별 OHLC 데이터를
    시가/고가/저가/종가/거래량 탭으로 나누어 저장.
    - 행: 종목
    - 열: 일자
    market="KR"이면 코드 6자리, "US"면 그대로.
    """
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

    # 모든 날짜 수집
    all_dates = set()
    for stock_data in data_list:
        if stock_data['history']:
            for daily in stock_data['history']:
                all_dates.add(daily['date'])

    sorted_dates = sorted(list(all_dates))
    if not sorted_dates:
        print("\n❌ 저장할 데이터가 없습니다.")
        return

    sheet_configs = [
        ('시가', 'open'),
        ('고가', 'high'),
        ('저가', 'low'),
        ('종가', 'close'),
        ('거래량', 'volume')
    ]

    for sheet_name, field_name in sheet_configs:
        # 기존 시트 여부
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            existing_dates = []
            for col in range(3, sheet.max_column + 1):
                val = sheet.cell(row=1, column=col).value
                try:
                    existing_dates.append(int(val))
                except Exception:
                    continue

            existing_data = {}
            for row in range(2, sheet.max_row + 1):
                name = sheet.cell(row=row, column=1).value
                code = sheet.cell(row=row, column=2).value
                if not name or not code:
                    continue
                code_str = str(code).strip()
                if market == "KR":
                    code_key = code_str.zfill(6)
                else:
                    code_key = code_str

                values = {}
                for col_idx, date_int in enumerate(existing_dates, 3):
                    values[str(date_int)] = sheet.cell(row=row, column=col_idx).value
                existing_data[code_key] = {'name': name, 'values': values}
        else:
            sheet = wb.create_sheet(sheet_name)
            existing_dates = []
            existing_data = {}

        merged_dates = set(existing_dates)
        for stock_data in data_list:
            if stock_data['history']:
                for daily in stock_data['history']:
                    try:
                        merged_dates.add(int(daily['date']))
                    except Exception:
                        continue

        sorted_dates_all = sorted(list(merged_dates))

        # 헤더
        sheet.cell(row=1, column=1, value='종목명')
        sheet.cell(row=1, column=2, value='종목코드')
        for col_idx, date_int in enumerate(sorted_dates_all, 3):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = date_int
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        for col_idx in (1, 2):
            cell = sheet.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        # 종목 코드 전체 집합
        all_codes = set(existing_data.keys())
        for stock_data in data_list:
            all_codes.add(stock_data['code'])

        for row_idx, code in enumerate(sorted(all_codes), start=2):
            # 이름
            if code in existing_data:
                name = existing_data[code]['name']
            else:
                # data_list에서 찾기
                name = next((s['name'] for s in data_list if s['code'] == code), code)

            sheet.cell(row=row_idx, column=1, value=name)
            sheet.cell(row=row_idx, column=2, value=code)

            # 기존 값
            values = existing_data.get(code, {}).get('values', {})

            # 신규 값
            new_values = {}
            stock_hist = next((s for s in data_list if s['code'] == code), None)
            if stock_hist and stock_hist['history']:
                for daily in stock_hist['history']:
                    try:
                        new_values[str(int(daily['date']))] = daily[field_name]
                    except Exception:
                        continue

            # 날짜별로 값 입력
            for col_idx, date_int in enumerate(sorted_dates_all, 3):
                key = str(date_int)
                val = new_values.get(key, values.get(key, ''))
                sheet.cell(row=row_idx, column=col_idx, value=val)

        # 열 너비
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 14
        for col_idx in range(3, len(sorted_dates_all) + 3):
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = 12

    wb.save(filename)
    print(f"\n✅ 엑셀 파일 저장 완료: {filename}")


def save_netbuy_to_excel(data_list, filename, market="KR"):
    """
    종목별 일별 투자자 순매수 수량 저장.
    - 시트: 순매수_개인 / 순매수_외국인 / 순매수_기관계
    - 행: 종목
    - 열: 날짜(YYYYMMDD)
    """
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    all_dates = set()
    for stock_data in data_list:
        for daily in stock_data.get("netbuy_history", []) or []:
            d = daily.get("date")
            if d:
                all_dates.add(d)

    if not all_dates:
        print("\n❌ 순매수 저장할 데이터가 없습니다.")
        return

    sorted_dates = sorted(all_dates)

    for key, sheet_name in NETBUY_SHEET_BY_KEY.items():
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            existing_dates = []
            for col in range(3, sheet.max_column + 1):
                val = sheet.cell(row=1, column=col).value
                digits = "".join(ch for ch in str(val) if ch.isdigit())
                if len(digits) == 8:
                    existing_dates.append(digits)

            existing_data = {}
            for row in range(2, sheet.max_row + 1):
                name = sheet.cell(row=row, column=1).value
                code = sheet.cell(row=row, column=2).value
                if not name or not code:
                    continue
                code_str = str(code).strip()
                if market == "KR":
                    code_key = code_str.zfill(6)
                else:
                    code_key = code_str
                values = {}
                for col_idx, date_str in enumerate(existing_dates, 3):
                    values[date_str] = sheet.cell(row=row, column=col_idx).value
                existing_data[code_key] = {"name": name, "values": values}
        else:
            sheet = wb.create_sheet(sheet_name)
            existing_dates = []
            existing_data = {}

        merged_dates = sorted(set(existing_dates) | set(sorted_dates))

        # 헤더
        sheet.cell(row=1, column=1, value="종목명")
        sheet.cell(row=1, column=2, value="종목코드")
        for col_idx, date_str in enumerate(merged_dates, 3):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = int(date_str)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        for col_idx in (1, 2):
            cell = sheet.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        all_codes = set(existing_data.keys())
        for stock_data in data_list:
            all_codes.add(stock_data["code"])

        for row_idx, code in enumerate(sorted(all_codes), start=2):
            name = existing_data.get(code, {}).get("name")
            if not name:
                name = next((s["name"] for s in data_list if s["code"] == code), code)

            sheet.cell(row=row_idx, column=1, value=name)
            sheet.cell(row=row_idx, column=2, value=code)

            old_values = existing_data.get(code, {}).get("values", {})
            new_values = {}
            stock_hist = next((s for s in data_list if s["code"] == code), None)
            if stock_hist:
                for daily in stock_hist.get("netbuy_history", []) or []:
                    d = daily.get("date")
                    if not d:
                        continue
                    new_values[d] = daily.get(key)

            for col_idx, date_str in enumerate(merged_dates, 3):
                val = new_values.get(date_str, old_values.get(date_str, ""))
                sheet.cell(row=row_idx, column=col_idx, value=val)

        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 14
        for col_idx in range(3, len(merged_dates) + 3):
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = 12

    wb.save(filename)
    print(f"\n✅ 순매수 시트 저장 완료: {filename}")


def get_latest_date_from_sheet(filename, sheet_name):
    """지정 시트(종가/거래량 등)에서 가장 최신 날짜를 'YYYYMMDD' 문자열로 반환"""
    try:
        wb = openpyxl.load_workbook(filename)
        if sheet_name not in wb.sheetnames:
            return None
        sheet = wb[sheet_name]

        dates = [sheet.cell(row=1, column=col).value for col in range(3, sheet.max_column + 1)]
        dates_dt = []
        for d in dates:
            try:
                dates_dt.append(datetime.strptime(str(d), '%Y%m%d'))
            except Exception:
                pass
        if not dates_dt:
            return None
        latest = max(dates_dt)
        return latest.strftime('%Y%m%d')
    except Exception as e:
        print(f"❌ 날짜 추출 에러({filename}/{sheet_name}): {e}")
        return None


# =========================
# 3. 지수(코스피/코스닥) 시트
# =========================

def fetch_index_history(access_token, domain, index_code, app_key, app_secret,
                        start_date, end_date):
    """
    업종지수 기간별 시세 조회 (일별)
    - index_code: 0001(KOSPI), 1001(KOSDAQ), 2001(KOSPI200)
    """
    endpoint = f"{domain}/uapi/domestic-stock/v1/quotations/inquire-daily-indexchartprice"

    params = {
        "fid_cond_mrkt_div_code": "U",
        "fid_input_iscd": index_code,
        "fid_input_date_1": start_date,
        "fid_input_date_2": end_date,
        "fid_period_div_code": "D",
    }

    headers = {
        "content-type": "application/json; charset=utf-8",
        "authorization": f"Bearer {access_token}",
        "appkey": app_key,
        "appsecret": app_secret,
        "tr_id": "FHKUP03500100",
        "custtype": "P",
    }

    try:
        resp = requests.get(endpoint, headers=headers, params=params, timeout=10)

        if resp.status_code != 200:
            print(f"❌ 업종지수 HTTP {resp.status_code} 오류 ({index_code})")
            print(resp.text)
            return None

        data = resp.json()
        rows = data.get("output2")

        if not rows:
            print(f"❌ 업종지수 데이터 없음 ({index_code})")
            return None

        history = []
        for row in rows:
            history.append({
                "date": row.get("stck_bsop_date", ""),
                "index_value": row.get("bstp_nmix_prpr"),
                "open": row.get("bstp_nmix_oprc"),
                "high": row.get("bstp_nmix_hgpr"),
                "low": row.get("bstp_nmix_lwpr"),
            })

        history.sort(key=lambda x: x["date"])
        return history

    except Exception as e:
        print(f"❌ 업종지수 조회 중 에러 ({index_code}): {e}")
        return None


def update_index_sheet(access_token, domain, app_key, app_secret,
                       filename="KR_Stocks_Individual.xlsx"):
    """
    파일의 '지수' 시트 업데이트
    - 없으면: 최근 100일치 KOSPI/KOSDAQ/KOSPI200 생성
    - 있으면: 마지막 날짜 이후 ~ 오늘까지 추가
    """
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

    indices = [
        ("KOSPI", "0001"),
        ("KOSDAQ", "1001"),
        ("KOSPI200", "2001"),
    ]

    def normalize_index_code(code):
        code_str = str(code).strip()
        return code_str.zfill(6) if code_str.isdigit() else code_str

    today = datetime.now()
    today_str = today.strftime('%Y%m%d')

    # A. 시트 없는 경우
    if '지수' not in wb.sheetnames:
        sheet = wb.create_sheet('지수')

        end_date = today_str
        start_date = (today - timedelta(days=100)).strftime('%Y%m%d')
        print(f"\n📈 [지수] 최초 생성: {start_date} ~ {end_date}")

        index_data = {}
        all_dates = set()

        for name, code in indices:
            print(f"  ▶ {name} ({code}) 조회 중...")
            history = fetch_index_history(
                access_token, domain, code, app_key, app_secret,
                start_date, end_date
            )
            if not history:
                print(f"    • {name} 데이터 없음")
                continue

            values = {}
            for h in history:
                d = h["date"]
                v = h["index_value"]
                if not d or v is None:
                    continue
                values[d] = float(v)
                all_dates.add(d)

            norm_code = normalize_index_code(code)
            index_data[norm_code] = {
                "name": name,
                "code": norm_code,
                "values": values
            }
            print(f"    • {len(values)}일치 데이터 확보")

            time.sleep(0.5)

        if not index_data or not all_dates:
            print("\n❌ 지수 데이터가 없어 저장하지 않습니다.")
            wb.save(filename)
            return

        sorted_dates = sorted(all_dates)

        # 헤더
        sheet.cell(row=1, column=1, value='업종명')
        sheet.cell(row=1, column=2, value='업종코드')
        sheet.cell(row=1, column=1).font = Font(bold=True)
        sheet.cell(row=1, column=2).font = Font(bold=True)
        sheet.cell(row=1, column=1).fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        sheet.cell(row=1, column=2).fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        for col_idx, date_str in enumerate(sorted_dates, 3):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = date_str
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        # 데이터
        for row_idx, code in enumerate(sorted(index_data.keys()), start=2):
            info = index_data[code]
            sheet.cell(row=row_idx, column=1, value=info["name"])
            sheet.cell(row=row_idx, column=2, value=info["code"])

            values = info["values"]
            for col_idx, date_str in enumerate(sorted_dates, 3):
                val = values.get(date_str, "")
                sheet.cell(row=row_idx, column=col_idx, value=val)

        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 12
        for col_idx in range(3, len(sorted_dates) + 3):
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = 12

        wb.save(filename)
        print(f"\n✅ '지수' 시트 최초 생성 완료: {filename}")
        return

    # B. 시트 있는 경우 → 추가
    sheet = wb['지수']
    print("\n📈 [지수] 기존 시트 업데이트 시작")

    existing_dates = []
    for col in range(3, sheet.max_column + 1):
        val = sheet.cell(row=1, column=col).value
        if val:
            existing_dates.append(str(val))

    existing_data = {}
    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(row=row, column=1).value
        code = sheet.cell(row=row, column=2).value
        if not code:
            continue
        code_str = normalize_index_code(code)
        values = {}
        for idx, date_str in enumerate(existing_dates, 3):
            values[date_str] = sheet.cell(row=row, column=idx).value
        existing_data[code_str] = {"name": name, "values": values}

    latest = get_latest_date_from_sheet(filename, "지수")
    if latest:
        start_dt = datetime.strptime(latest, "%Y%m%d") + timedelta(days=1)
        start_date = start_dt.strftime("%Y%m%d")
        print(f"  • 마지막 날짜: {latest} → 추가 조회 시작일: {start_date}")
    else:
        start_date = (today - timedelta(days=100)).strftime("%Y%m%d")
        print(f"  • 기존 날짜 없음 → {start_date} ~ {today_str} 재조회")

    end_date = today_str
    if datetime.strptime(start_date, "%Y%m%d") > datetime.strptime(end_date, "%Y%m%d"):
        print("  • 추가할 지수 데이터가 없습니다. (이미 최신)")
        return

    new_index_data = {}
    all_dates = set(existing_dates)

    for name, code in indices:
        print(f"  ▶ {name} ({code}) 신규 조회: {start_date} ~ {end_date}")
        history = fetch_index_history(
            access_token, domain, code, app_key, app_secret,
            start_date, end_date
        )
        if not history:
            print(f"    • {name} 추가 데이터 없음")
            continue

        values = {}
        for h in history:
            d = h["date"]
            v = h["index_value"]
            if not d or v is None:
                continue
            values[d] = float(v)
            all_dates.add(d)

        norm_code = normalize_index_code(code)
        new_index_data[norm_code] = {
            "name": name,
            "code": norm_code,
            "values": values
        }
        print(f"    • {len(values)}일치 신규 데이터 확보")

        time.sleep(0.5)

    if not new_index_data:
        print("  • 신규 지수 데이터가 없어 업데이트하지 않습니다.")
        return

    merged_dates = sorted(all_dates)

    # 헤더
    sheet.cell(row=1, column=1, value='업종명')
    sheet.cell(row=1, column=2, value='업종코드')
    sheet.cell(row=1, column=1).font = Font(bold=True)
    sheet.cell(row=1, column=2).font = Font(bold=True)
    sheet.cell(row=1, column=1).fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    sheet.cell(row=1, column=2).fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    for col_idx, date_str in enumerate(merged_dates, 3):
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = date_str
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    all_codes = set(existing_data.keys()) | set(normalize_index_code(code) for _, code in indices)

    for row_idx, code in enumerate(sorted(all_codes), start=2):
        if code in existing_data:
            name = existing_data[code]["name"]
        else:
            name = next((n for (n, c) in indices if normalize_index_code(c) == code), code)

        sheet.cell(row=row_idx, column=1, value=name)
        sheet.cell(row=row_idx, column=2, value=code)

        old_values = existing_data.get(code, {}).get("values", {})
        new_values = new_index_data.get(code, {}).get("values", {})

        for col_idx, date_str in enumerate(merged_dates, 3):
            val = new_values.get(date_str, old_values.get(date_str, ""))
            sheet.cell(row=row_idx, column=col_idx, value=val)

    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 12
    for col_idx in range(3, len(merged_dates) + 3):
        col_letter = get_column_letter(col_idx)
        sheet.column_dimensions[col_letter].width = 12

    wb.save(filename)
    print(f"\n✅ '지수' 시트 업데이트 완료: {filename}")


# =========================
# 4. 카테고리별 처리 래퍼
# =========================

def fetch_kr_wrapper(access_token, domain, symbol, start_date, end_date,
                     app_key, app_secret):
    """국내(개별+ETF) 모두 동일 API 사용"""
    return fetch_stock_daily_history(
        access_token, domain, symbol, start_date, end_date,
        app_key=app_key, app_secret=app_secret
    )


def fetch_us_wrapper(access_token, domain, symbol, start_date, end_date,
                     app_key, app_secret):
    """
    미국(개별+ETF) 모두 해외 기간별 시세 API 사용
    - 여기서는 일단 NASDAQ("NAS") 기준으로 호출
      (나중에 필요하면 엑셀에 EXCD 컬럼 추가해서 확장)
    """
    return fetch_overseas_daily_history(
        access_token, domain, market_code="NAS", symbol=symbol,
        start_date=start_date, end_date=end_date,
        app_key=app_key, app_secret=app_secret
    )


def process_one_file(excel_filename, fetch_func, app_key, app_secret, domain,
                     access_token, market="KR", update_index=False, update_netbuy=False):
    """
    엑셀 파일 1개 처리:
    - 종목 목록 로드
    - 종가/거래량 시트 기준으로 날짜 범위 결정
    - fetch_func으로 각 종목 히스토리 가져오기
    - 시가/고가/저가/종가/거래량 시트 저장
    - 필요 시 지수 시트 업데이트
    """
    stocks = load_stock_list(excel_filename, market=market)
    if not stocks:
        return

    latest_close = get_latest_date_from_sheet(excel_filename, "종가")
    latest_amount = get_latest_date_from_sheet(excel_filename, "거래량")

    today = datetime.now()
    today_str = today.strftime('%Y%m%d')

    if latest_close and latest_amount:
        latest_str = max(latest_close, latest_amount)
        start_dt = datetime.strptime(latest_str, '%Y%m%d') + timedelta(days=1)
        start_date = start_dt.strftime('%Y%m%d')
        print(f"\n📅 [{excel_filename}] 추가 조회: {start_date} ~ {today_str}")
    else:
        end_dt = today
        start_date = (end_dt - timedelta(days=100)).strftime('%Y%m%d')
        print(f"\n📅 [{excel_filename}] 전체 조회(최근 100일): {start_date} ~ {today_str}")

    end_date = today_str

    netbuy_start_date = None
    if update_netbuy:
        latest_netbuy = []
        for sheet_name in NETBUY_SHEET_BY_KEY.values():
            d = get_latest_date_from_sheet(excel_filename, sheet_name)
            if d:
                latest_netbuy.append(d)
        if latest_netbuy:
            latest_netbuy_str = max(latest_netbuy)
            netbuy_start_dt = datetime.strptime(latest_netbuy_str, "%Y%m%d") + timedelta(days=1)
            netbuy_start_date = netbuy_start_dt.strftime("%Y%m%d")
            print(f"📅 [{excel_filename}] 순매수 추가 조회: {netbuy_start_date} ~ {today_str}")
        else:
            netbuy_start_date = start_date
            print(f"📅 [{excel_filename}] 순매수 전체 조회(기준): {netbuy_start_date} ~ {today_str}")

    data_list = []
    netbuy_data_list = []
    print(f"\n총 {len(stocks)}개 종목에 대해 조회합니다... ({excel_filename})")
    for i, stock in enumerate(stocks, start=1):
        print(f"  [{i}/{len(stocks)}] {stock['name']}({stock['code']}) ...", end='')

        history = fetch_func(
            access_token, domain,
            stock['code'], start_date, end_date,
            app_key, app_secret
        )

        if history:
            print(f"성공 ({len(history)}일)")
            data_list.append({
                "name": stock['name'],
                "code": stock['code'],
                "history": history,
            })
        else:
            print("실패 또는 추가 데이터 없음")

        if update_netbuy and netbuy_start_date and netbuy_start_date <= today_str:
            netbuy_history = fetch_investor_netbuy_history(
                access_token=access_token,
                domain=domain,
                symbol=stock["code"],
                start_date=netbuy_start_date,
                app_key=app_key,
                app_secret=app_secret,
            )
            if netbuy_history:
                netbuy_data_list.append({
                    "name": stock["name"],
                    "code": stock["code"],
                    "netbuy_history": netbuy_history,
                })

        time.sleep(1)

    if data_list:
        save_history_to_excel(data_list, filename=excel_filename, market=market)
        if update_index and market == "KR":
            update_index_sheet(
                access_token=access_token,
                domain=domain,
                app_key=app_key,
                app_secret=app_secret,
                filename=excel_filename
            )
    else:
        print(f"\n❌ [{excel_filename}] 저장할 데이터가 없습니다.")

    if update_netbuy:
        if netbuy_start_date and netbuy_start_date > today_str:
            print(f"  • [{excel_filename}] 순매수는 이미 최신입니다.")
        elif netbuy_data_list:
            save_netbuy_to_excel(netbuy_data_list, filename=excel_filename, market=market)
        else:
            print(f"  • [{excel_filename}] 순매수 신규 데이터가 없습니다.")


# =========================
# 5. main
# =========================

def main():
    print(f"\n=== 한국투자증권 API 주식/ETF 시세 히스토리 조회 ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')}) ===")

    secrets = load_api_secrets()
    if not secrets:
        return

    app_key = secrets.get('api_key')
    app_secret = secrets.get('api_secret')
    domain = secrets.get('domain', 'https://openapi.koreainvestment.com:9443')

    file_config = load_file_config('stock_file_map.json')
    if not file_config:
        return

    print("\n🔄 토큰 발급 요청 중...")
    token_data = get_token(app_key, app_secret, domain)
    if not token_data:
        print("\n❌ 토큰 발급 실패")
        return
    access_token = token_data['access_token']

    # 카테고리별 설정
    # - market: 코드 처리(KR: zfill, US: 그대로)
    # - fetch_func: 어떤 API 호출할지
    # - update_index: 지수 시트 생성/업데이트 여부
    category_settings = {
        "KR_Stocks_Individual": {"market": "KR", "fetch_func": fetch_kr_wrapper, "update_index": True, "update_netbuy": True},
        "KR_Stocks_ETF":        {"market": "KR", "fetch_func": fetch_kr_wrapper, "update_index": False, "update_netbuy": False},
        "US_Stocks_Individual": {"market": "US", "fetch_func": fetch_us_wrapper, "update_index": False, "update_netbuy": False},
        "US_Stocks_ETF":        {"market": "US", "fetch_func": fetch_us_wrapper, "update_index": False, "update_netbuy": False},
    }

    for category_name, excel_filename in file_config.items():
        print("\n=======================================")
        print(f"📂 카테고리: {category_name}")
        print(f"📊 파일: {excel_filename}")
        print("=======================================")

        cfg = category_settings.get(category_name)
        if not cfg:
            print(f"⚠️ {category_name} 에 대한 설정이 없습니다. 건너뜀.")
            continue

        process_one_file(
            excel_filename=excel_filename,
            fetch_func=cfg["fetch_func"],
            app_key=app_key,
            app_secret=app_secret,
            domain=domain,
            access_token=access_token,
            market=cfg["market"],
            update_index=cfg["update_index"],
            update_netbuy=cfg.get("update_netbuy", False),
        )


if __name__ == "__main__":
    main()
