"""
File: recompute_std_last10.py
Version: v1.1.0
Role: 최근 10일간 STD 값을 새로운 산식으로 재계산해 원본 파일을 덮어쓴다.
"""

from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from zipfile import BadZipFile
import shutil
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP

# helper
def normalize_code(value):
    if pd.isna(value):
        return None
    s = str(value).strip()
    if not s:
        return None
    digits = "".join(ch for ch in s if ch.isdigit())
    if len(digits) == 0:
        return s
    return digits.zfill(6)

BASE_DIR = Path(__file__).resolve().parent  # 스크립트가 있는 폴더
# 대상 파일 경로 후보 ((std) 붙은 파일만 탐색)
EXCEL_CANDIDATES = [
    BASE_DIR / "KR_Stocks_ETF.xlsx",        # 스크립트와 같은 폴더
    BASE_DIR.parent / "KR_Stocks_ETF.xlsx", # 상위(리포 루트)
]
STD_SHEET = "std"
PRICE_SHEET = "종가"
WINDOW = 20
RECENT_DAYS = 20
START_DATE_LABEL = 20251226  # YYYYMMDD 기준 재계산 시작 지점


def load_std_sheet(path: Path):
    df = pd.read_excel(path, sheet_name=STD_SHEET, dtype={"종목코드": str})
    if "종목코드" in df.columns:
        df["종목코드"] = df["종목코드"].apply(normalize_code)
    if df.empty or len(df.columns) < 3:
        raise ValueError(f"{STD_SHEET} 시트에 데이터가 없습니다.")
    return df


def load_close_prices(path: Path):
    df = pd.read_excel(path, sheet_name=PRICE_SHEET, dtype={"종목코드": str})
    if df.empty or len(df.columns) < 3:
        raise ValueError(f"{PRICE_SHEET} 시트에 데이터가 없습니다.")

    df["종목코드"] = df["종목코드"].apply(normalize_code)
    date_cols = [c for c in df.columns if c not in ("종목명", "종목코드")]
    if not date_cols:
        raise ValueError(f"{PRICE_SHEET} 시트에 날짜 열이 없습니다.")

    norm_labels = []
    rename_map = {}
    for col in date_cols:
        lbl = format_date_label(col)
        norm_labels.append(lbl)
        rename_map[col] = lbl

    df = df.rename(columns=rename_map)
    numeric_df = df[norm_labels].apply(pd.to_numeric, errors="coerce")

    price_matrix = {}
    for idx, code in enumerate(df["종목코드"].tolist()):
        price_matrix[code] = numeric_df.iloc[idx].to_numpy(dtype=float)

    return price_matrix, norm_labels


def recompute_std(values: np.ndarray) -> float:
    """
    extra_scores.py의 calc_std_value와 동일한 산식:
    - idx 시점(마지막 값)에서 window_std=20 롤링 σ 계산
    - 과거 window_mean=20 동안의 σ 평균 대비 증감률 * 100
    - 소수 둘째 자리 반올림
    """
    window_std = WINDOW
    window_mean = WINDOW
    min_len = window_std + window_mean - 1
    if len(values) < min_len:
        return np.nan

    idx = len(values) - 1
    min_idx = window_std + window_mean - 2
    if idx < min_idx:
        return np.nan

    std_list = []
    for j in range(idx - window_mean + 1, idx + 1):  # j: idx-19 ~ idx
        start = j - window_std + 1
        end = j + 1
        if start < 0:
            return np.nan

        window_prices = values[start:end]
        if any(pd.isna(window_prices)):
            return np.nan

        arr = np.array(window_prices, dtype=float)
        sigma = float(np.std(arr, ddof=0))  # 모표준편차
        std_list.append(sigma)

    if not std_list:
        return np.nan

    std_today = std_list[-1]
    avg_std = sum(std_list) / len(std_list)
    if avg_std == 0:
        return 0.0

    raw_val = (std_today / avg_std - 1) * 100
    val = Decimal(str(raw_val)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return float(val)


def update_recent_std(df: pd.DataFrame, date_columns, price_matrix, label_positions, start_label=None, verbose=False):
    normalized_labels = [format_date_label(col) for col in date_columns]

    start_index = max(0, len(date_columns) - RECENT_DAYS)
    if start_label:
        for idx, lbl in enumerate(normalized_labels):
            try:
                lbl_int = int(lbl)
            except (TypeError, ValueError):
                continue
            if lbl_int >= start_label:
                start_index = idx
                break

    codes = df["종목코드"].tolist()
    if verbose:
        target_cols = date_columns[start_index:]
        print(f"▶ STD 재계산 시작 (start_label={start_label}, start_index={start_index}, 대상 컬럼 수={len(target_cols)})")

    skipped_cols = 0
    for idx_col in range(start_index, len(date_columns)):
        raw_label = date_columns[idx_col]
        norm_label = normalized_labels[idx_col]
        pos = label_positions.get(norm_label)
        if pos is None:
            if verbose:
                print(f"  - 스킵: {raw_label} (정규 라벨 {norm_label} 위치를 찾을 수 없음)")
            skipped_cols += 1
            continue

        # 기존 데이터를 비운 후 새 값으로 채움
        df[raw_label] = np.nan

        new_values = []
        for code in codes:
            series = price_matrix.get(code)
            if series is None or pos >= len(series):
                new_values.append(np.nan)
                continue
            std_val = recompute_std(series[:pos + 1])
            new_values.append(std_val)

        new_values = np.array(new_values, dtype=float)
        mask = ~np.isnan(new_values)
        new_values[mask] = np.round(new_values[mask], 2)
        df[raw_label] = new_values

        if verbose:
            print(
                f"  - 완료: {raw_label} (pos={pos}) "
                f"재계산 {int(mask.sum())}건 / NaN {int(len(new_values) - mask.sum())}건"
            )

    if verbose:
        print(f"▶ 재계산 완료: 처리 컬럼 {len(date_columns) - start_index - skipped_cols}, 스킵 {skipped_cols}")

    return df


def format_date_label(value):
    if isinstance(value, (datetime, pd.Timestamp)):
        return int(value.strftime("%Y%m%d"))

    s = str(value).strip()
    if not s:
        return s

    digits = "".join(ch for ch in s if ch.isdigit())
    if len(digits) == 8 and digits.isdigit():
        return int(digits)

    s = s.replace(".", "").replace("-", "").replace("/", "")
    digits = "".join(ch for ch in s if ch.isdigit())
    if len(digits) == 8 and digits.isdigit():
        return int(digits)

    return s


def normalize_date_columns(df: pd.DataFrame) -> pd.DataFrame:
    renamed = {}
    for col in df.columns:
        if col in ("종목명", "종목코드"):
            renamed[col] = col
        else:
            renamed[col] = format_date_label(col)
    return df.rename(columns=renamed)


def align_std_columns(df: pd.DataFrame, close_labels):
    """
    STD 시트 컬럼을 종가 시트 날짜에 맞춰 확장/정렬한다.
    부족한 날짜 컬럼이 있으면 NaN으로 추가하고, 종가 순서대로 재배치한다.
    """
    std_cols = [c for c in df.columns if c not in ("종목명", "종목코드")]
    need_add = [lbl for lbl in close_labels if lbl not in std_cols]
    for lbl in need_add:
        df[lbl] = np.nan
    ordered_cols = ["종목명", "종목코드"] + close_labels
    return df[ordered_cols], close_labels


def overwrite_std_sheet(path: Path, df: pd.DataFrame):
    """openpyxl로 STD 시트를 안전하게 덮어쓰기"""
    try:
        wb = load_workbook(path)
    except BadZipFile as e:
        raise RuntimeError(f"엑셀 파일이 손상되었습니다: {e}")

    # 기존 STD 시트 제거
    if STD_SHEET in wb.sheetnames:
        ws_old = wb[STD_SHEET]
        idx = wb.sheetnames.index(STD_SHEET)
        wb.remove(ws_old)
    else:
        idx = len(wb.sheetnames)

    ws = wb.create_sheet(STD_SHEET, idx)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    wb.save(path)
    wb.close()


def main():
    excel_path = None
    for cand in EXCEL_CANDIDATES:
        if cand.exists():
            excel_path = cand
            break
    if excel_path is None:
        raise FileNotFoundError(f"KR_Stocks_ETF.xlsx 파일을 찾을 수 없습니다. (확인한 경로: {EXCEL_CANDIDATES})")
    print(f"📄 사용 파일: {excel_path}")

    # STD 시트의 날짜 헤더를 정규화(YYYYMMDD)해서 가격 시트와 매칭되도록 보정
    df = normalize_date_columns(load_std_sheet(excel_path))
    price_matrix, close_labels = load_close_prices(excel_path)
    label_positions = {lbl: idx for idx, lbl in enumerate(close_labels)}
    std_date_cols = [c for c in df.columns if c not in ("종목명", "종목코드")]
    print(f"📅 STD 컬럼 수: {len(std_date_cols)}, 종가 날짜 수: {len(close_labels)}")
    if std_date_cols:
        print(f"   STD 날짜 범위: {std_date_cols[0]} ~ {std_date_cols[-1]}")
    if close_labels:
        print(f"   종가 날짜 범위: {close_labels[0]} ~ {close_labels[-1]}")

    # STD 컬럼을 종가 날짜에 맞춰 확장/정렬
    df, date_cols = align_std_columns(df, close_labels)
    if len(date_cols) < WINDOW:
        raise ValueError(f"날짜 열이 {WINDOW}개 이상 필요합니다. (현재 {len(date_cols)})")

    updated = update_recent_std(
        df.copy(),
        date_cols,
        price_matrix,
        label_positions,
        start_label=START_DATE_LABEL,
        verbose=True,
    )
    updated = normalize_date_columns(updated)
    if "종목코드" in updated.columns:
        updated["종목코드"] = updated["종목코드"].apply(normalize_code)

    backup = excel_path.with_suffix(".bak")
    shutil.copy2(excel_path, backup)
    try:
        overwrite_std_sheet(excel_path, updated)
    except Exception:
        shutil.move(backup, excel_path)
        raise
    else:
        if backup.exists():
            backup.unlink()

    print(f"✅ STD 시트 최신 {RECENT_DAYS}일 갱신 완료: {excel_path.name}")


if __name__ == "__main__":
    main()
