"""
File: set_cell_value.py
Version: v1.0.0
Role: 지정한 엑셀 파일의 특정 시트·셀 값을 변경한다.

# 사용법
- 아래 상수(EXCEL_PATH, SHEET_NAME, CELL_ADDR, NEW_VALUE)를 필요에 맞게 수정한 뒤 실행:
    python set_cell_value.py
- 변경 전에 파일이 존재하는지와 시트명이 올바른지 확인하세요.
"""

from pathlib import Path
import openpyxl

BASE_DIR = Path(__file__).resolve().parent

# ===== 변경할 대상 설정 (필요에 맞게 수정) =====
# 엑셀 파일이 상위 폴더(리포 루트)에 있는 경우 기본값을 부모 경로로 설정
EXCEL_PATH = BASE_DIR.parent / "KR_Stocks_ETF.xlsx"
SHEET_NAME = "std"                       # 시트 이름
CELL_ADDR = "B2"                         # 변경할 셀 주소 (예: "C3")
NEW_VALUE = "0000Z0"                      # 셀에 쓸 값 (문자/숫자/날짜 등)
# VALUE_MODE: "auto" → 있는 그대로, "str" → 문자열로 강제, "int" → 정수로 강제, "float" → 실수로 강제
VALUE_MODE = "auto"
# ============================================


def convert_value(value, mode: str):
    mode = (mode or "auto").lower()
    if mode == "str":
        return str(value)
    if mode == "int":
        try:
            return int(value)
        except Exception:
            raise ValueError(f"VALUE_MODE=int 이지만 int 변환 실패: {value}")
    if mode == "float":
        try:
            return float(value)
        except Exception:
            raise ValueError(f"VALUE_MODE=float 이지만 float 변환 실패: {value}")
    # auto: 그대로 사용
    return value


def set_cell_value(file_path: Path, sheet_name: str, cell_addr: str, new_value):
    if not file_path.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path}")

    wb = openpyxl.load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"시트를 찾을 수 없습니다: {sheet_name}")

    ws = wb[sheet_name]
    value_to_write = convert_value(new_value, VALUE_MODE)
    ws[cell_addr] = value_to_write
    wb.save(file_path)
    wb.close()
    print(f"✅ {file_path.name} [{sheet_name}] {cell_addr} -> {value_to_write} (mode={VALUE_MODE})")


def main():
    set_cell_value(EXCEL_PATH, SHEET_NAME, CELL_ADDR, NEW_VALUE)


if __name__ == "__main__":
    main()
