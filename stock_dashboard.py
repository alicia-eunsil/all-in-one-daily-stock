"""
File: stock_dashboard.py
Version: v4.2.5
Role: 계산된 주가 지표와 원자료를 조회하는 Streamlit 대시보드.
# 메모: v4.0.1 - GAP/STD 상하위 강조 시 평균·지수 행 제외 + 동률 포함, 불필요 포맷 함수 제거
# 메모: v4.0.2 - Z30 출력 추가
# 메모: v4.0.3 - 대시보드 표시에서 Z30 숨김(데이터 로딩은 유지)
# 메모: v4.1.0 - GAP 시트는 GAP20으로 표시하고 GAP60(=gap60) 지표를 추가
# 메모: v4.1.1 - 데이터프레임 높이를 브라우저 높이(100vh) 기반으로 동적 조정
# 메모: v4.2.0 - KR_Stocks_Individual에 개인/외국인/기관계 순매수(일별) '매수량' 탭 추가
# 메모: v4.2.1 - sigmat/isigmat는 화면에서 20일 표준편차 의미인 STD20으로 표시
# 메모: v4.2.2 - 접속코드 bcrypt 해시는 소스 하드코딩 대신 환경변수/Secrets에서 읽도록 변경
# 메모: v4.2.3 - 접속코드는 평문 환경변수/Secrets ACCESS_CODE로도 읽을 수 있게 지원
# 메모: v4.2.4 - STD20을 종합/지표별에서 분리해 별도 '표준편차' 탭으로 표시
# 메모: v4.2.5 - 표준편차 탭 강조 규칙 적용 + 지표별/원자료/매수량/표준편차 기본 표시일을 20일로 확대
"""

import streamlit as st
import subprocess
import sys
import os
import gc
import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime, date, timedelta
import json  # 🔥 4개 엑셀 매핑용

# ======================================
# 페이지 설정 (최초 UI 출력 전에 호출)
# ======================================
st.set_page_config(page_title="주식 데이터 대시보드", page_icon="🚀", layout="wide", initial_sidebar_state="collapsed")

# ======================================
# 0. 인증 (간단 비밀번호)
# ======================================
def load_access_code():
    # 운영에서는 평문 ACCESS_CODE를 우선 사용하고, 없으면 기존 bcrypt 해시도 fallback 지원한다.
    plain_value = os.getenv("ACCESS_CODE")
    if not plain_value:
        try:
            plain_value = st.secrets.get("ACCESS_CODE")
        except Exception:
            plain_value = None

    if plain_value:
        return {"mode": "plain", "value": str(plain_value)}

    hash_value = os.getenv("ACCESS_CODE_HASH")
    if not hash_value:
        try:
            hash_value = st.secrets.get("ACCESS_CODE_HASH")
        except Exception:
            hash_value = None

    if hash_value:
        return {"mode": "hash", "value": str(hash_value).encode()}

    return None


ACCESS_CODE_CONFIG = load_access_code()

if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False

if ACCESS_CODE_CONFIG is None:
    st.error("ACCESS_CODE 또는 ACCESS_CODE_HASH 환경변수/Secrets가 설정되지 않았습니다.")
    st.stop()

if not st.session_state["authenticated"]:
    st.title("🔒 Access Required")
    st.write("Please enter the access code to open the dashboard.")

    with st.form("auth_form"):
        code = st.text_input("Enter access code", type="password")
        submitted = st.form_submit_button("Submit")

    if submitted:
        is_valid = False
        if ACCESS_CODE_CONFIG["mode"] == "plain":
            is_valid = code == ACCESS_CODE_CONFIG["value"]
        else:
            import bcrypt
            is_valid = bcrypt.checkpw(code.encode(), ACCESS_CODE_CONFIG["value"])

        if is_valid:
            st.session_state["authenticated"] = True
            st.success("Access granted")
            st.rerun()
        else:
            st.error("Invalid code")

    st.stop()

# ======================================
# 1. 전역 상태 변수
# ======================================
if "run_update" not in st.session_state:
    st.session_state.run_update = False
if "data_loaded" not in st.session_state:
    st.session_state.data_loaded = True

# 🔥 종합 탭 날짜 확장용
if "show_days" not in st.session_state:
    st.session_state.show_days = 20  # 시작: 최근 20일

# 🔥 원자료 탭 날짜 확장용
if "show_days_raw" not in st.session_state:
    st.session_state.show_days_raw = 20  # 시작: 최근 20일
if "show_days_buy" not in st.session_state:
    st.session_state.show_days_buy = 20  # 시작: 최근 20일
if "show_days_std20" not in st.session_state:
    st.session_state.show_days_std20 = 20  # 시작: 최근 20일

# 🔥 파일 선택 상태
BASE_DIR = Path(__file__).resolve().parent
JSON_PATH = BASE_DIR / "stock_file_map.json"
EXCEL_MAP = {}
INDEX_SHEET_NAME = "지수"
NETBUY_SHEET_MAP = {
    "개인": "순매수_개인",
    "외국인": "순매수_외국인",
    "기관계": "순매수_기관계",
}
if "selected_category" not in st.session_state:
    # JSON 로드해서 첫 번째 항목을 기본 선택값으로
    try:
        with open(JSON_PATH, "r", encoding="utf-8") as f:
            _tmp_map = json.load(f)
        if isinstance(_tmp_map, dict) and _tmp_map:
            st.session_state.selected_category = list(_tmp_map.keys())[0]
        else:
            st.session_state.selected_category = None
    except:
        st.session_state.selected_category = None


# ======================================
# 2. 날짜/포맷 유틸 함수
# ======================================
def _to_datetime(v):
    """엑셀/문자열/숫자 등 다양한 형태의 날짜를 datetime으로 변환"""
    if isinstance(v, (datetime, date)):
        return datetime(v.year, v.month, v.day)

    if isinstance(v, (int, float)):
        iv = int(v)
        digits = str(iv)
        if len(digits) == 8 and digits.isdigit():
            try:
                return datetime.strptime(digits, "%Y%m%d")
            except:
                pass
        base = datetime(1899, 12, 30)
        try:
            return base + timedelta(days=iv)
        except:
            return None

    s = str(v).strip()
    if not s:
        return None

    for fmt in ("%Y-%m-%d", "%Y.%m.%d.", "%Y.%m.%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt)
        except:
            pass

    digits = "".join(ch for ch in s if ch.isdigit())
    if len(digits) == 8:
        try:
            return datetime.strptime(digits, "%Y%m%d")
        except:
            pass

    return None


def format_excel_date(v):
    """_to_datetime로 바꾼 날짜를 YYYY.MM.DD. 형식 문자열로 변환"""
    dt = _to_datetime(v)
    if dt:
        return dt.strftime("%Y.%m.%d.")
    s = str(v)
    s = s.replace("-", ".").replace("/", ".")
    if not s.endswith("."):
        s += "."
    return s


def _inject_responsive_dataframe_height_css():
    """
    브라우저 높이 기준으로 데이터프레임 높이를 동적으로 조정한다.
    - 상단 타이틀/필터 영역을 고려해 오프셋(330px)을 뺀다.
    - 너무 작거나 큰 화면에서 과도해지지 않게 min/max를 둔다.
    """
    st.markdown(
        """
        <style>
        :root {
            --dynamic-table-height: min(1400px, max(520px, calc(100vh - 330px)));
        }
        div[data-testid="stDataFrame"] {
            height: var(--dynamic-table-height) !important;
        }
        div[data-testid="stDataFrame"] > div {
            height: 100% !important;
        }
        div[data-testid="stDataFrame"] [data-testid="stVirtualizedDataFrame"] {
            height: 100% !important;
        }
        div[data-testid="stDataFrame"] * {
            scrollbar-width: auto;
            scrollbar-color: #9ca3af #f3f4f6;
        }
        div[data-testid="stDataFrame"] *::-webkit-scrollbar {
            width: 14px;
            height: 14px;
        }
        div[data-testid="stDataFrame"] *::-webkit-scrollbar-track {
            background: #f3f4f6;
        }
        div[data-testid="stDataFrame"] *::-webkit-scrollbar-thumb {
            background-color: #9ca3af;
            border: 3px solid #f3f4f6;
            border-radius: 8px;
        }
        div[data-testid="stDataFrame"] *::-webkit-scrollbar-thumb:hover {
            background-color: #6b7280;
        }
        div[data-testid="stDataFrame"] *::-webkit-scrollbar-corner {
            background: #f3f4f6;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


_inject_responsive_dataframe_height_css()


def _highlight_top_bottom_cells(df_or_styler, columns, top_n=10, high_color="#ffe0e0", low_color="#e0ecff", allowed_mask=None):
    """
    각 컬럼에서 상위/하위 N개에 배경색을 칠한 Styler 반환.
    컬럼 값이 문자열이어도 to_numeric으로 비교.
    동률(같은 값)은 cutoff 값과 같으면 모두 포함한다.
    """
    if not columns:
        return df_or_styler

    styler = df_or_styler if hasattr(df_or_styler, "apply") else df_or_styler.style

    def _style(col):
        styles = [""] * len(col)
        series = pd.to_numeric(col, errors="coerce")
        if allowed_mask is not None:
            mask_use = allowed_mask.reindex(col.index, fill_value=False)
        else:
            mask_use = pd.Series(True, index=col.index)
        valid = series[mask_use].dropna()
        if valid.empty:
            return styles
        n_use = min(top_n, len(valid))
        top_vals = valid.nlargest(n_use)
        bottom_vals = valid.nsmallest(n_use)
        top_cut = top_vals.min()
        bottom_cut = bottom_vals.max()
        top_idx = set(valid[valid >= top_cut].index)
        bottom_idx = set(valid[valid <= bottom_cut].index)
        for i, idx in enumerate(col.index):
            if idx in top_idx:
                styles[i] = f"background-color: {high_color}; color: red"
            elif idx in bottom_idx:
                styles[i] = f"background-color: {low_color}; color: blue"
        return styles

    for col in columns:
        styler = styler.apply(_style, subset=pd.IndexSlice[:, col])
    return styler


def _highlight_row_min_max_cells(df_or_styler, columns, lookback_n=20, high_color="#ffe0e0", low_color="#e0ecff", allowed_mask=None):
    """
    각 행(종목)에서 최근 lookback_n개 컬럼 값 중 최댓값/최솟값 셀을 강조.
    동률은 모두 포함한다.
    """
    if not columns:
        return df_or_styler

    use_cols = list(columns[-lookback_n:])
    if not use_cols:
        return df_or_styler

    styler = df_or_styler if hasattr(df_or_styler, "apply") else df_or_styler.style

    def _style_row(row):
        styles = [""] * len(row)

        if allowed_mask is not None:
            allowed = bool(allowed_mask.reindex([row.name], fill_value=False).iloc[0])
            if not allowed:
                return styles

        series = pd.to_numeric(row, errors="coerce")
        valid = series.dropna()
        if valid.empty:
            return styles

        max_val = valid.max()
        min_val = valid.min()
        for i, num in enumerate(series):
            if pd.isna(num):
                continue
            if num == max_val:
                styles[i] = f"background-color: {high_color}; color: red"
            elif num == min_val:
                styles[i] = f"background-color: {low_color}; color: blue"
        return styles

    return styler.apply(_style_row, axis=1, subset=use_cols)


def _highlight_threshold(df_or_styler, columns, high_cond, low_cond, high_color="#ffe0e0", low_color="#e0ecff"):
    """
    조건부(상/하)로 색을 입힌 Styler 반환.
    """
    if not columns:
        return df_or_styler

    styler = df_or_styler if hasattr(df_or_styler, "apply") else df_or_styler.style

    def _style(col):
        styles = [""] * len(col)
        for i, val in enumerate(col):
            try:
                num = float(val)
            except:
                continue
            if high_cond(num):
                styles[i] = f"background-color: {high_color}; color: red"
            elif low_cond(num):
                styles[i] = f"background-color: {low_color}; color: blue"
        return styles

    for col in columns:
        styler = styler.apply(_style, subset=pd.IndexSlice[:, col])
    return styler


def _load_index_metric_data(wb, selected_labels):
    sheet_map = {
        "S20": "is20",
        "S60": "is60",
        "S120": "is120",
        "Z20": "iz20",
        "Z30": "iz30",
        "Z60": "iz60",
        "Z120": "iz120",
        "GAP20": "igap",
        "GAP60": "igap60",
        "STD": "istd",
        # 저장 시트명은 isigmat를 유지하되, 화면에서는 의미 기준으로 STD20으로 노출한다.
        "STD20": "isigmat",
    }
    combined = {}
    per_metric = {}
    for metric, sheet_name in sheet_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        label_to_col = {}
        for col in range(3, ws.max_column + 1):
            raw = ws.cell(row=1, column=col).value
            if raw is None:
                continue
            lbl = format_excel_date(raw)
            label_to_col[lbl] = col

        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=1).value
            code = ws.cell(row=row, column=2).value
            if not name or not code:
                continue
            code_str = str(code).strip()
            if code_str.isdigit():
                code_str = code_str.zfill(6)
            entry = combined.setdefault(code_str, {"code": code_str, "name": str(name), "values": {}})
            metric_row = {"code": code_str, "name": str(name), "values": {}}
            for lbl in selected_labels:
                col_idx = label_to_col.get(lbl)
                val = ws.cell(row=row, column=col_idx).value if col_idx else None
                entry["values"][(lbl, metric)] = val
                metric_row["values"][lbl] = val
            per_metric.setdefault(metric, []).append(metric_row)

    combined_list = list(combined.values()) if combined else []
    return combined_list, per_metric

# ======================================
# 3. 뷰 렌더링 함수들
# ======================================
def render_total_view(indicator_df, selected_labels, indicator_range_msg, total_days,
                     index_df=None, index_metric_rows=None):
    """
    1️⃣ 종합 탭
    - 멀티헤더(날짜×지표) 구조
    - 맨 아래 평균 행
    - 그 아래 KOSPI/KOSDAQ/KOSPI200 행 추가
    """
    if indicator_df is None:
        st.warning("⚠️ 종합 데이터를 불러올 수 없습니다.")
        return

    st.markdown("### 🔍 필터 옵션 (종합)")
    c1, c2 = st.columns(2)
    with c1:
        search = st.text_input("🔎 종목명/종목코드 검색", key="search_total")
    with c2:
        sort_by = st.selectbox("정렬 기준", ["종목코드", "종목명"], key="sort_total")

    # 검색 적용
    df_f = indicator_df.copy()
    if search:
        df_f = df_f[
            df_f["종목명"].astype(str).str.contains(search, case=False) |
            df_f["종목코드"].astype(str).str.contains(search, case=False)
        ]

    df_f = df_f.sort_values(by=sort_by)

    st.info(indicator_range_msg)

    # --------------------------------------
    # 🔥 멀티헤더 생성 (1행: 날짜, 2행: 지표명)
    # --------------------------------------
    metrics = ["Z20", "Z60", "Z120", "S20", "S60", "S120", "GAP20", "GAP60", "QUANT", "STD"]
    col_tuples = [("", "종목코드"), ("", "종목명")]
    df_show_data = {
        ("", "종목코드"): df_f["종목코드"].to_numpy(copy=False),
        ("", "종목명"): df_f["종목명"].to_numpy(copy=False),
    }

    # 날짜 × 지표 조합을 한 번에 생성해 DataFrame 파편화와 메모리 급증을 막는다.
    for lbl in selected_labels:
        for m in metrics:
            key = (lbl, m)
            if key in df_f.columns:
                df_show_data[(lbl, m)] = df_f[key].to_numpy(copy=False)
            else:
                df_show_data[(lbl, m)] = ["-"] * len(df_f)
            col_tuples.append((lbl, m))

    df_show = pd.DataFrame(df_show_data).reset_index(drop=True)
    df_show.columns = pd.MultiIndex.from_tuples(col_tuples)

    # --------------------------------------
    # 🔥 평균 행 추가 (맨 마지막 행)
    # --------------------------------------
    avg_row = []
    for col in df_show.columns:
        if col == ("", "종목코드"):
            avg_row.append("AVG")
        elif col == ("", "종목명"):
            avg_row.append("평균")
        else:
            lbl, m = col
            key = (lbl, m)
            if key in df_f.columns:
                s = pd.to_numeric(df_f[key], errors="coerce")
                avg_val = s.mean(skipna=True)
                avg_row.append(f"{avg_val:.2f}")
            else:
                avg_row.append(None)

    df_show.loc[len(df_show)] = avg_row  # 평균 행 추가

    index_rows_df = None
    if index_metric_rows:
        index_records = []
        for row in index_metric_rows:
            row_dict = {("", "종목코드"): row["code"], ("", "종목명"): row["name"]}
            for (lbl, metric), val in row["values"].items():
                col = (lbl, metric)
                if col in df_show.columns:
                    row_dict[col] = val
            index_records.append(row_dict)

        if index_records:
            index_rows_df = pd.DataFrame(index_records)
            index_rows_df = index_rows_df.reindex(columns=df_show.columns)
            df_show = pd.concat([df_show, index_rows_df], ignore_index=True)

            avg_idx_vals = []
            for col in df_show.columns:
                if col == ("", "종목코드"):
                    avg_idx_vals.append("INDEX AVG")
                elif col == ("", "종목명"):
                    avg_idx_vals.append("지수 평균")
                else:
                    series = pd.to_numeric(index_rows_df[col], errors="coerce")
                    mean_val = series.mean(skipna=True)
                    avg_idx_vals.append(mean_val if pd.notna(mean_val) else None)
            df_show.loc[len(df_show)] = avg_idx_vals

    # --------------------------------------
    # Z/S/Q/GAP20/GAP60 포맷 적용
    # --------------------------------------
    for lbl in selected_labels:
        for m in ["Z20", "Z60", "Z120", "S20", "S60", "S120", "GAP20", "GAP60", "STD", "QUANT"]:
            col = (lbl, m)
            if col in df_show.columns:
                df_show[col] = pd.to_numeric(df_show[col], errors="coerce")

    z_columns_total = [(lbl, m) for lbl in selected_labels for m in ["Z20", "Z60", "Z120"] if (lbl, m) in df_show.columns]
    s_columns_total = [(lbl, m) for lbl in selected_labels for m in ["S20", "S60", "S120"] if (lbl, m) in df_show.columns]
    gap20_columns_total = [(lbl, "GAP20") for lbl in selected_labels if (lbl, "GAP20") in df_show.columns]
    gap60_columns_total = [(lbl, "GAP60") for lbl in selected_labels if (lbl, "GAP60") in df_show.columns]
    std_columns_total = [(lbl, "STD") for lbl in selected_labels if (lbl, "STD") in df_show.columns]
    quant_columns_total = [(lbl, "QUANT") for lbl in selected_labels if (lbl, "QUANT") in df_show.columns]

    # --------------------------------------
    # 🔽 지수(KOSPI/KOSDAQ/KOSPI200) 행 추가 (레거시, 지수 지표 데이터 없을 때)
    # --------------------------------------
    if (not index_metric_rows) and index_df is not None and not index_df.empty:
        for _, idx_row in index_df.iterrows():
            new_row_vals = []
            used_dates = set()

            for col in df_show.columns:
                if col == ("", "종목코드"):
                    new_row_vals.append(idx_row.get("업종코드", ""))
                elif col == ("", "종목명"):
                    new_row_vals.append(idx_row.get("업종명", ""))
                else:
                    lbl, m = col
                    if lbl not in used_dates:
                        val = idx_row.get(lbl, None)
                        new_row_vals.append(val if pd.notna(val) else "")
                        used_dates.add(lbl)
                    else:
                        new_row_vals.append("")

            df_show.loc[len(df_show)] = new_row_vals

    # 인덱스 설정 (종목코드·종목명)
    df_show = df_show.set_index([("", "종목코드"), ("", "종목명")])

    # 숫자 포맷 적용
    fmt_map = {}
    for col in (z_columns_total + s_columns_total + gap20_columns_total + gap60_columns_total + quant_columns_total):
        fmt_map[col] = "{:.0f}"
    for col in std_columns_total:
        fmt_map[col] = "{:.2f}"

    styler = df_show.style.format(fmt_map, na_rep="-")

    # 강조 색상 적용 (AVG / 지수 행 제외)
    disallow_codes = {"AVG", "INDEX AVG"}
    if index_rows_df is not None and not index_rows_df.empty:
        disallow_codes.update(index_rows_df[("", "종목코드")].astype(str))

    mask_codes = ~df_show.index.get_level_values(0).astype(str).isin(disallow_codes)
    allowed_mask_total = pd.Series(mask_codes, index=df_show.index)

    # 강조 색상 적용
    if gap20_columns_total:
        styler = _highlight_row_min_max_cells(styler, gap20_columns_total, lookback_n=20, allowed_mask=allowed_mask_total)
    if gap60_columns_total:
        styler = _highlight_row_min_max_cells(styler, gap60_columns_total, lookback_n=20, allowed_mask=allowed_mask_total)
    if std_columns_total:
        styler = _highlight_row_min_max_cells(styler, std_columns_total, lookback_n=20, allowed_mask=allowed_mask_total)
    if z_columns_total:
        styler = _highlight_threshold(styler, z_columns_total,
                                      high_cond=lambda v: v > 100,
                                      low_cond=lambda v: v < -100)
    if s_columns_total:
        styler = _highlight_threshold(styler, s_columns_total,
                                      high_cond=lambda v: abs(v - 100) < 0.1,
                                      low_cond=lambda v: abs(v - 0) < 0.1)
    if quant_columns_total:
        styler = _highlight_threshold(styler, quant_columns_total,
                                      high_cond=lambda v: v > 100,
                                      low_cond=lambda v: v < 25)

    st.dataframe(
    styler,
    use_container_width=True,
)

    # 🔥 과거 확장 버튼 (종합)
    if st.button("⬅ 과거 10일 더보기(종합)", disabled=(total_days <= st.session_state.show_days)):
        st.session_state.show_days = min(st.session_state.show_days + 10, total_days)
        st.rerun()


def render_metric_view(indicator_df, selected_labels, index_metric_map=None):
    """
    2️⃣ 지표별 탭:
    - 1열: 종목코드
    - 2열: 종목명
    - 이후: 날짜별 선택 지표값
    """
    st.subheader("📈 지표 선택")

    if indicator_df is None or len(indicator_df) == 0:
        st.warning("⚠️ 지표별 데이터를 불러올 수 없습니다.")
        return

    metric_options = ["Z20", "Z60", "Z120",
                      "S20", "S60", "S120",
                      "GAP20", "GAP60", "QUANT", "STD"]

    # 실제 존재하는 지표만
    available = []
    for m in metric_options:
        if any(((lbl, m) in indicator_df.columns) for lbl in selected_labels):
            available.append(m)

    if not available:
        st.error("indicator_df에 S/Z/GAP20/GAP60/QUANT/STD 관련 컬럼이 없습니다.")
        st.write("현재 indicator_df.columns 예시:", list(indicator_df.columns)[:20])
        return

    metric = st.selectbox("지표를 선택하세요", available, index=0)

    # -------------------------
    # DF 구성 (종목코드, 종목명 + 날짜별 값)
    # -------------------------
    df_metric_numeric = indicator_df[["종목코드", "종목명"]].copy()

    for lbl in selected_labels:
        col_key = (lbl, metric)
        if col_key in indicator_df.columns:
            df_metric_numeric[lbl] = indicator_df[col_key]
        else:
            df_metric_numeric[lbl] = None

    # 숫자 변환
    for lbl in selected_labels:
        if lbl in df_metric_numeric.columns:
            df_metric_numeric[lbl] = pd.to_numeric(df_metric_numeric[lbl], errors="coerce")

    # 🔍 필터 + 정렬
    st.markdown("### 🔍 필터 옵션 (지표별)")
    c1, c2 = st.columns(2)
    with c1:
        search_metric = st.text_input("🔎 종목명/종목코드 검색", key="search_metric")
    with c2:
        sort_metric = st.selectbox("정렬 기준", ["종목코드", "종목명"], key="sort_metric")

    df_filtered_display = df_metric_numeric.copy()
    df_filtered_numeric = df_metric_numeric.copy()
    if search_metric:
        mask = (
            df_filtered_display["종목명"].astype(str).str.contains(search_metric, case=False)
            | df_filtered_display["종목코드"].astype(str).str.contains(search_metric, case=False)
        )
        df_filtered_display = df_filtered_display[mask]
        df_filtered_numeric = df_filtered_numeric[mask]

    sort_order = df_filtered_display.sort_values(by=sort_metric).index
    df_filtered_display = df_filtered_display.loc[sort_order].reset_index(drop=True)
    df_filtered_numeric = df_filtered_numeric.loc[sort_order].reset_index(drop=True)

    gap_columns_metric = []
    std_columns_metric = []
    if metric in ("GAP20", "GAP60"):
        gap_columns_metric = [lbl for lbl in selected_labels if lbl in df_filtered_display.columns]
    elif metric == "STD":
        std_columns_metric = [lbl for lbl in selected_labels if lbl in df_filtered_display.columns]

    # 평균 행 추가
    avg_row = {"종목코드": "AVG", "종목명": "평균"}
    for lbl in selected_labels:
        if lbl not in df_filtered_numeric.columns:
            continue
        col_numeric = pd.to_numeric(df_filtered_numeric[lbl], errors="coerce")
        mean_val = col_numeric.mean(skipna=True)
        avg_row[lbl] = mean_val if pd.notna(mean_val) else None

    df_filtered = df_filtered_display.copy()
    df_filtered.loc[len(df_filtered)] = avg_row

    idx_rows = []
    if index_metric_map and metric in index_metric_map:
        idx_rows = index_metric_map[metric]
        if idx_rows:
            idx_display_records = []
            for row in idx_rows:
                rec = {"종목코드": row["code"], "종목명": row["name"]}
                for lbl in selected_labels:
                    rec[lbl] = pd.to_numeric(row["values"].get(lbl), errors="coerce")
                idx_display_records.append(rec)

            if idx_display_records:
                df_idx_display = pd.DataFrame(idx_display_records)
                df_filtered = pd.concat([df_filtered, df_idx_display], ignore_index=True)

                avg_idx_row = {"종목코드": "INDEX AVG", "종목명": "지수 평균"}
                for lbl in selected_labels:
                    vals = [row["values"].get(lbl) for row in idx_rows]
                    series = pd.Series(pd.to_numeric(vals, errors="coerce"))
                    mean_val = series.mean(skipna=True)
                    avg_idx_row[lbl] = mean_val
                df_filtered.loc[len(df_filtered)] = avg_idx_row

    # 날짜 범위 안내
    if selected_labels:
        oldest_label = selected_labels[0]
        latest_label = selected_labels[-1]
        st.info(
            f"📅 지표별 표시 범위: **{oldest_label} ~ {latest_label}** "
            f"(최근 {len(selected_labels)}일)"
        )

    # 테이블 출력
    st.markdown(f"### 📋 {metric} · 추이")

    column_config = {
        "종목코드": st.column_config.TextColumn("종목코드", width="small", pinned="left"),
        "종목명": st.column_config.TextColumn("종목명", width="small", pinned="left"),
    }
    for lbl in selected_labels:
        if lbl in df_filtered.columns:
            column_config[lbl] = st.column_config.NumberColumn(lbl, format="%.2f" if metric == "STD" else "%.0f")

    # 강조에서 평균/지수 행 제외
    base_len = len(df_filtered_display)  # 순수 종목 행 개수
    idx_count = len(idx_rows)
    avg_row_idx = base_len  # 종목 평균 행
    idx_start = base_len + 1
    idx_end = base_len + idx_count
    idx_avg_idx = base_len + idx_count + 1

    allowed_mask = pd.Series(True, index=df_filtered.index)
    if avg_row_idx < len(allowed_mask):
        allowed_mask.iloc[avg_row_idx] = False
    if idx_count:
        allowed_mask.iloc[idx_start:idx_end + 1] = False
        if idx_avg_idx < len(allowed_mask):
            allowed_mask.iloc[idx_avg_idx] = False

    # 스타일 (색상 강조)
    metric_fmt = "{:.2f}" if metric == "STD" else "{:.0f}"
    styler = df_filtered.style.format({lbl: metric_fmt for lbl in selected_labels}, na_rep="-")
    if metric.startswith("GAP"):
        styler = _highlight_row_min_max_cells(styler, gap_columns_metric, lookback_n=20, allowed_mask=allowed_mask)
    elif metric == "STD":
        styler = _highlight_row_min_max_cells(styler, std_columns_metric, lookback_n=20, allowed_mask=allowed_mask)
    elif metric.startswith("Z"):
        styler = _highlight_threshold(styler, [lbl for lbl in selected_labels],
                                      high_cond=lambda v: v > 100, low_cond=lambda v: v < -100)
    elif metric.startswith("S"):
        styler = _highlight_threshold(styler, [lbl for lbl in selected_labels],
                                      high_cond=lambda v: abs(v - 100) < 0.1, low_cond=lambda v: abs(v - 0) < 0.1)
    elif metric == "QUANT":
        styler = _highlight_threshold(styler, [lbl for lbl in selected_labels],
                                      high_cond=lambda v: v > 100, low_cond=lambda v: v < 25)

    st.dataframe(
        styler,
        use_container_width=True,
        hide_index=True,
        column_config=column_config,
    )

    # 🔥 과거 확장 버튼 (지표별)
    global total_days
    if st.button("⬅ 과거 10일 더보기(지표별)", disabled=(total_days <= st.session_state.show_days)):
        st.session_state.show_days = min(st.session_state.show_days + 10, total_days)
        st.rerun()


def render_raw_view(close_df, close_range_msg, total_close_days, index_close_rows=None):
    """
    3️⃣ 원자료(종가) 탭
    - 종목코드/종목명 + 날짜별 종가
    """
    if close_df is None:
        st.warning("⚠️ 원자료(종가) 데이터를 불러올 수 없습니다.")
        return

    st.markdown("### 🔍 필터 옵션 (원자료)")
    r1, r2 = st.columns(2)
    with r1:
        search_raw = st.text_input("🔎 종목명/종목코드 검색", key="search_raw")
    with r2:
        sort_raw = st.selectbox("정렬 기준", ["종목코드", "종목명"], key="sort_raw")

    df_raw = close_df.copy()

    if search_raw:
        df_raw = df_raw[
            df_raw["종목코드"].astype(str).str.contains(search_raw, case=False) |
            df_raw["종목명"].astype(str).str.contains(search_raw, case=False)
        ]

    df_raw = df_raw.sort_values(by=sort_raw)

    st.info(close_range_msg)

    # 날짜 컬럼 추출
    date_cols = [c for c in df_raw.columns if c not in ["종목코드", "종목명"]]

    # 컬럼 순서 고정
    df_raw = df_raw[["종목코드", "종목명"] + date_cols]

    if index_close_rows:
        for row in index_close_rows:
            new_row = {"종목코드": row.get("업종코드", ""), "종목명": row.get("업종명", "")}
            for c in date_cols:
                new_row[c] = row.get(c, None)
            df_raw.loc[len(df_raw)] = new_row

    column_config = {
        "종목코드": st.column_config.TextColumn("종목코드", width="small", pinned="left"),
        "종목명": st.column_config.TextColumn("종목명", width="small", pinned="left"),
    }
    for c in date_cols:
        column_config[c] = st.column_config.NumberColumn(c, width="small")

    for c in date_cols:
        df_raw[c] = pd.to_numeric(df_raw[c], errors="coerce")

    st.dataframe(
        df_raw.style.format({c: "{:,.0f}" for c in date_cols}, na_rep="-"),
        use_container_width=True,
        hide_index=True,
        column_config=column_config,
    )

    # 🔥 과거 확장 버튼 (원자료)
    if st.button("⬅ 과거 10일 더보기(종가)", disabled=(total_close_days <= st.session_state.show_days_raw)):
        st.session_state.show_days_raw = min(st.session_state.show_days_raw + 10, total_close_days)
        st.rerun()


def render_netbuy_view(netbuy_df_map, netbuy_range_msg, total_netbuy_days):
    """
    4️⃣ 매수량 탭
    - 종목코드/종목명 + (날짜 × 투자자유형) 순매수 수량
    """
    if not netbuy_df_map:
        st.warning("⚠️ 매수량 데이터를 불러올 수 없습니다.")
        return

    investor_order = [name for name in ["개인", "외국인", "기관계"] if name in netbuy_df_map]
    if not investor_order:
        st.warning("⚠️ 매수량 데이터의 투자자 유형을 찾을 수 없습니다.")
        return

    base_df = netbuy_df_map[investor_order[0]].copy()

    st.markdown("### 🔍 필터 옵션 (매수량)")
    b1, b2 = st.columns(2)
    with b1:
        search_buy = st.text_input("🔎 종목명/종목코드 검색", key="search_buy")
    with b2:
        sort_buy = st.selectbox("정렬 기준", ["종목코드", "종목명"], key="sort_buy")

    if search_buy:
        base_df = base_df[
            base_df["종목코드"].astype(str).str.contains(search_buy, case=False) |
            base_df["종목명"].astype(str).str.contains(search_buy, case=False)
        ]

    base_df = base_df.sort_values(by=sort_buy)
    st.info(netbuy_range_msg)

    date_cols = [c for c in base_df.columns if c not in ["종목코드", "종목명"]]
    if not date_cols:
        st.warning("⚠️ 매수량 날짜 컬럼이 없습니다.")
        return

    row_keys = pd.MultiIndex.from_frame(base_df[["종목코드", "종목명"]])
    part_frames = []
    for investor in investor_order:
        df_inv = netbuy_df_map[investor].copy()
        df_inv = df_inv.set_index(["종목코드", "종목명"])
        df_inv = df_inv.reindex(row_keys)
        df_inv = df_inv.reindex(columns=date_cols)
        df_inv.columns = pd.MultiIndex.from_tuples([(d, investor) for d in date_cols])
        part_frames.append(df_inv)

    df_show = pd.concat(part_frames, axis=1)

    ordered_cols = []
    for d in date_cols:
        for investor in investor_order:
            col = (d, investor)
            if col in df_show.columns:
                ordered_cols.append(col)
    df_show = df_show.reindex(columns=ordered_cols)

    st.dataframe(
        df_show.style.format("{:,.0f}", na_rep="-"),
        use_container_width=True,
    )

    if st.button("⬅ 과거 10일 더보기(매수량)", disabled=(total_netbuy_days <= st.session_state.show_days_buy)):
        st.session_state.show_days_buy = min(st.session_state.show_days_buy + 10, total_netbuy_days)
        st.rerun()


def render_std20_view(std20_df, std20_range_msg, total_std20_days, index_std20_rows=None):
    """
    5️⃣ 표준편차 탭
    - 종목코드/종목명 + 날짜별 STD20
    """
    if std20_df is None:
        st.warning("⚠️ 표준편차(STD20) 데이터를 불러올 수 없습니다.")
        return

    st.markdown("### 🔍 필터 옵션 (표준편차)")
    s1, s2 = st.columns(2)
    with s1:
        search_std20 = st.text_input("🔎 종목명/종목코드 검색", key="search_std20")
    with s2:
        sort_std20 = st.selectbox("정렬 기준", ["종목코드", "종목명"], key="sort_std20")

    df_std20 = std20_df.copy()

    if search_std20:
        df_std20 = df_std20[
            df_std20["종목코드"].astype(str).str.contains(search_std20, case=False) |
            df_std20["종목명"].astype(str).str.contains(search_std20, case=False)
        ]

    df_std20 = df_std20.sort_values(by=sort_std20)
    st.info(std20_range_msg)

    date_cols = [c for c in df_std20.columns if c not in ["종목코드", "종목명"]]
    if not date_cols:
        st.warning("⚠️ 표준편차(STD20) 날짜 컬럼이 없습니다.")
        return

    df_std20 = df_std20[["종목코드", "종목명"] + date_cols]

    if index_std20_rows:
        for row in index_std20_rows:
            new_row = {"종목코드": row.get("업종코드", ""), "종목명": row.get("업종명", "")}
            for c in date_cols:
                new_row[c] = row.get(c, None)
            df_std20.loc[len(df_std20)] = new_row

    column_config = {
        "종목코드": st.column_config.TextColumn("종목코드", width="small", pinned="left"),
        "종목명": st.column_config.TextColumn("종목명", width="small", pinned="left"),
    }
    for c in date_cols:
        column_config[c] = st.column_config.NumberColumn(c, width="small")

    for c in date_cols:
        df_std20[c] = pd.to_numeric(df_std20[c], errors="coerce")

    # STD20은 실제 표준편차 절대값이라 큰 수가 나올 수 있어 천단위 콤마를 표시한다.
    styler = df_std20.style.format({c: "{:,.2f}" for c in date_cols}, na_rep="-")

    # 표준편차 탭은 STD와 같은 규칙을 적용한다:
    # 최근 20일 중 각 종목 행의 최댓값은 빨강, 최솟값은 파랑.
    disallow_codes = set()
    if index_std20_rows:
        disallow_codes.update(str(row.get("업종코드", "")).strip() for row in index_std20_rows)

    allowed_mask = ~df_std20["종목코드"].astype(str).isin(disallow_codes)
    allowed_mask.index = df_std20.index
    styler = _highlight_row_min_max_cells(styler, date_cols, lookback_n=20, allowed_mask=allowed_mask)

    st.dataframe(
        styler,
        use_container_width=True,
        hide_index=True,
        column_config=column_config,
    )

    if st.button("⬅ 과거 10일 더보기(표준편차)", disabled=(total_std20_days <= st.session_state.show_days_std20)):
        st.session_state.show_days_std20 = min(st.session_state.show_days_std20 + 10, total_std20_days)
        st.rerun()


# ======================================
# 4. 엑셀 파일 매핑(JSON) 로드
# ======================================
try:
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        EXCEL_MAP = json.load(f)
    if not isinstance(EXCEL_MAP, dict) or not EXCEL_MAP:
        st.error("stock_file_map.json 형식이 잘못되었거나 비어 있습니다.")
        st.stop()
except FileNotFoundError:
    st.error(f"stock_file_map.json 파일을 찾을 수 없습니다. 위치: {JSON_PATH}")
    st.stop()
except Exception as e:
    st.error(f"stock_file_map.json 읽기 오류: {e}")
    st.stop()

categories = list(EXCEL_MAP.keys())
if not categories:
    st.error("stock_file_map.json 에 항목이 없습니다.")
    st.stop()
    sys.exit(1)  # streamlit 외 환경에서 실행될 때 안전 종료

if st.session_state.selected_category not in categories:
    st.session_state.selected_category = categories[0]

# ======================================
# 5. 상단: 네 개 카테고리 선택 버튼 (라디오, 가로)
# ======================================
st.markdown("### 📂 조회할 주식 그룹 선택")

selected_category = st.radio(
    "주식 그룹",
    categories,
    key="selected_category",
    horizontal=True,
    label_visibility="collapsed",
)
selected_filename = EXCEL_MAP[selected_category]
excel_path = (BASE_DIR / selected_filename).resolve()
is_index_target = excel_path.stem == "KR_Stocks_Individual"

#st.markdown(f"#### 현재: `{selected_category}`")

# ======================================
# 6. 사이드바: 선택 파일 다운로드 + 전체 갱신 버튼
# ======================================
with st.sidebar:
    # 🔄 최신 데이터 수동 새로고침 (git pull 후 캐시 초기화)
    if st.button("🔄 데이터 새로고침", key="refresh_git_pull"):
        try:
            subprocess.run(["git", "pull", "--rebase"], check=True)
        except Exception as e:
            st.error(f"git pull 실패: {e}")
        st.cache_data.clear()
        st.rerun()

    st.markdown("### 📁 현재 선택 파일")
    st.write(f"`{selected_filename}`")

    if excel_path.exists():
        with open(excel_path, "rb") as f:
            st.download_button(
                label="📥 선택 파일 다운로드",
                data=f,
                file_name=excel_path.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_excel",
            )
    else:
        st.warning(f"`{selected_filename}` 파일이 아직 생성되지 않았습니다.")

    st.markdown("---")
    if st.button("🔄 네 개 파일 전체 데이터 갱신"):
        st.session_state.run_update = True

# ======================================
# 7. 데이터 갱신 실행 (외부 스크립트 호출)
# ======================================
if st.session_state.run_update:
    with st.sidebar:
        st.subheader("진행 상황")
        status_box = st.empty()

    scripts = [
        ("run_all_scores.py", "4개 엑셀 S/Z + GAP20/GAP60/QUANT/STD 계산"),
    ]

    for idx, (sc, desc) in enumerate(scripts):
        status_box.info(f"{desc} 실행 중...")
        try:
            result = subprocess.run(
                [sys.executable, sc], capture_output=True, text=True, timeout=1800
            )
            if result.returncode == 0:
                st.sidebar.success(f"{desc} 완료")
            else:
                st.sidebar.error(f"{desc} 실패")
                st.sidebar.code(result.stderr[:500])
        except Exception as e:
                st.sidebar.error(f"{desc} 오류 발생: {e}")

        status_box.success(f"{desc} 완료")

    st.session_state.data_loaded = True
    st.session_state.run_update = False
    st.rerun()

# ======================================
# 8. 선택된 엑셀 파일 로드
# ======================================
if not excel_path.exists():
    st.error(f"`{selected_filename}` 파일을 찾지 못했습니다. "
             "왼쪽의 '네 개 파일 전체 데이터 갱신' 버튼으로 먼저 데이터를 생성해 주세요.")
    st.stop()

excel_file = excel_path
wb = openpyxl.load_workbook(excel_file, data_only=True)

# ======================================
# 9. 종목 정보 로딩 (종목 시트)
# ======================================
stock_info = {}
if "종목" in wb.sheetnames:
    ws = wb["종목"]
    for r in ws.iter_rows(min_row=2, max_col=2):
        name = r[0].value
        code = r[1].value
        if code and name:
            stock_info[code] = name

# ======================================
# 10. 종합(Z20/Z30/Z60/Z120/S/GAP20/GAP60/QUANT/STD) 데이터 로딩
# ======================================
sheet_names = ["z20", "z30", "z60", "z120", "s20", "s60", "s120", "gap", "gap60", "quant", "std", "sigmat"]
display_metric_by_sheet = {
    "gap": "GAP20",
    "gap60": "GAP60",
    # 저장 시트명 sigmat는 유지하지만, 화면과 내부 표시 키는 STD20으로 통일한다.
    "sigmat": "STD20",
}

base_ws = None
for s in sheet_names:
    if s in wb.sheetnames:
        base_ws = wb[s]
        break

indicator_df = None
indicator_date_infos = []
total_days = 0
selected_labels = []
indicator_range_msg = ""
index_metric_rows = []
index_metric_map = {}

if base_ws:
    max_col = base_ws.max_column

    # 기준 시트에서 날짜 헤더 수집 (1행, 3열~)
    for col in range(3, max_col + 1):
        raw = base_ws.cell(row=1, column=col).value
        if raw is None:
            continue
        dt = _to_datetime(raw)
        label = format_excel_date(raw)
        indicator_date_infos.append((col, raw, dt, label))

    indicator_date_infos = sorted(
        indicator_date_infos,
        key=lambda x: (x[2] is None, x[2] or datetime.min)
    )

    total_days = len(indicator_date_infos)

    show_days = min(st.session_state.show_days, total_days)
    start_idx = total_days - show_days
    selected_infos = indicator_date_infos[start_idx:]
    selected_labels = [lbl for _, _, _, lbl in selected_infos]

    oldest_label = selected_infos[0][3]
    latest_label = selected_infos[-1][3]
    indicator_range_msg = (
        f"📅 종합 표시 범위: **{oldest_label} ~ {latest_label}** "
        f"(최근 {show_days}일 / 전체 {total_days}일)"
    )

    # 종목별 데이터 딕셔너리 구성
    data_dict = {code: {"종목코드": code, "종목명": name}
                 for code, name in stock_info.items()}

    # 시트별 데이터 채우기 (날짜 문자열 기준 매칭)
    for s in sheet_names:
        if s not in wb.sheetnames:
            continue

        ws = wb[s]
        max_row_s = ws.max_row
        max_col_s = ws.max_column

        label_to_col = {}
        for col in range(3, max_col_s + 1):
            raw = ws.cell(row=1, column=col).value
            if raw is None:
                continue
            lbl = format_excel_date(raw)
            label_to_col[lbl] = col

        for r in range(2, max_row_s + 1):
            code = ws.cell(row=r, column=2).value
            if code not in data_dict:
                continue

            for lbl in selected_labels:
                col_idx = label_to_col.get(lbl)
                if col_idx is None:
                    val = None
                else:
                    val = ws.cell(row=r, column=col_idx).value

                metric_key = display_metric_by_sheet.get(s, s.upper())
                data_dict[code][(lbl, metric_key)] = val

    indicator_df = pd.DataFrame.from_dict(data_dict, orient="index").reset_index(drop=True)
    del data_dict

else:
    indicator_df = None

if indicator_df is not None and is_index_target and selected_labels:
    index_metric_rows, index_metric_map = _load_index_metric_data(wb, selected_labels)

# ======================================
# 11. 원자료(종가) 데이터 로딩
# ======================================
close_df = None
close_date_infos = []
total_close_days = 0
close_range_msg = ""
index_close_rows = []

if "종가" in wb.sheetnames:
    ws = wb["종가"]
    max_col_c = ws.max_column

    # 날짜 헤더
    for col in range(3, max_col_c + 1):
        raw = ws.cell(row=1, column=col).value
        if raw is None:
            continue

        dt = _to_datetime(raw)

        if dt is None:
            digits = "".join(ch for ch in str(raw) if ch.isdigit())
            if len(digits) == 8:
                dt = datetime.strptime(digits, "%Y%m%d")

        if dt is None:
            continue

        label = dt.strftime("%Y.%m.%d.")
        close_date_infos.append((col, raw, dt, label))

    close_date_infos = sorted(
        close_date_infos,
        key=lambda x: (x[2] is None, x[2] or datetime.min)
    )

    total_close_days = len(close_date_infos)

    show_raw = min(st.session_state.show_days_raw, total_close_days)
    start_idx = total_close_days - show_raw
    selected_close_infos = close_date_infos[start_idx:]

    oldest_label = selected_close_infos[0][3]
    latest_label = selected_close_infos[-1][3]

    close_range_msg = (
        f"📅 종가 표시 범위: **{oldest_label} ~ {latest_label}** "
        f"(최근 {show_raw}일 / 전체 {total_close_days}일)"
    )

    close_dict = {code: {"종목명": name, "종목코드": code}
                  for code, name in stock_info.items()}

    max_row_c = ws.max_row

    for r in range(2, max_row_c + 1):
        code = ws.cell(row=r, column=2).value
        if code not in close_dict:
            continue

        for col_idx, raw, dt, label in selected_close_infos:
            val = ws.cell(row=r, column=col_idx).value
            close_dict[code][label] = val

    close_df = pd.DataFrame.from_dict(close_dict, orient="index").reset_index(drop=True)
    del close_dict

    # 컬럼 이름을 yyyy.mm.dd. 형식으로 통일
    rename_map = {}
    for col in close_df.columns:
        if col in ["종목코드", "종목명"]:
            continue
        rename_map[col] = format_excel_date(col)

    close_df = close_df.rename(columns=rename_map)

    if is_index_target and INDEX_SHEET_NAME in wb.sheetnames:
        ws_idx_raw = wb[INDEX_SHEET_NAME]
        # 지수 시트의 날짜 헤더를 직접 매핑해 종가 시트와 동일한 라벨을 맞춘다.
        label_to_col_close = {}
        for col in range(3, ws_idx_raw.max_column + 1):
            raw = ws_idx_raw.cell(row=1, column=col).value
            if raw is None:
                continue
            lbl = format_excel_date(raw)
            label_to_col_close[lbl] = col

        rows_raw = []
        for r in range(2, ws_idx_raw.max_row + 1):
            name = ws_idx_raw.cell(row=r, column=1).value
            code = ws_idx_raw.cell(row=r, column=2).value
            if not name or not code:
                continue
            code_str = str(code).strip()
            if code_str.isdigit():
                code_str = code_str.zfill(6)
            row_dict = {"업종명": str(name), "업종코드": code_str}
            for lbl in close_df.columns:
                if lbl in ("종목코드", "종목명"):
                    continue
                col_idx = label_to_col_close.get(lbl)
                if col_idx is None:
                    row_dict[lbl] = None
                else:
                    row_dict[lbl] = ws_idx_raw.cell(row=r, column=col_idx).value
            rows_raw.append(row_dict)
        index_close_rows = rows_raw

# ======================================
# 12. 지수(KOSPI/KOSDAQ/KOSPI200) 데이터 로딩
# ======================================
index_df = None
netbuy_df_map = {}
total_netbuy_days = 0
netbuy_range_msg = ""
std20_df = None
index_std20_rows = []
total_std20_days = 0
std20_range_msg = ""

if "지수" in wb.sheetnames and indicator_df is not None and selected_labels:
    ws_idx = wb["지수"]
    max_col_i = ws_idx.max_column

    index_date_infos = []
    for col in range(3, max_col_i + 1):
        raw = ws_idx.cell(row=1, column=col).value
        if raw is None:
            continue

        dt = _to_datetime(raw)
        if dt is None:
            digits = "".join(ch for ch in str(raw) if ch.isdigit())
            if len(digits) == 8:
                dt = datetime.strptime(digits, "%Y%m%d")
        if dt is None:
            continue

        label = dt.strftime("%Y.%m.%d.")
        index_date_infos.append((col, raw, dt, label))

    label_to_col_idx = {label: col for col, raw, dt, label in index_date_infos}

    index_rows = []
    max_row_i = ws_idx.max_row

    seen_codes = set()
    for r in range(2, max_row_i + 1):
        name = ws_idx.cell(row=r, column=1).value
        code = ws_idx.cell(row=r, column=2).value
        if not name or not code:
            continue

        code_str = str(code).strip()
        if code_str.isdigit():
            code_str = code_str.zfill(6)
        if code_str in seen_codes:
            continue
        seen_codes.add(code_str)

        row_dict = {
            "업종명": str(name),
            "업종코드": code_str,
        }

        for lbl in selected_labels:
            col_idx = label_to_col_idx.get(lbl)
            if col_idx is None:
                val = None
            else:
                val = ws_idx.cell(row=r, column=col_idx).value
            row_dict[lbl] = val

        index_rows.append(row_dict)

    if index_rows:
        index_df = pd.DataFrame(index_rows)

# ======================================
# 12.3. 표준편차(STD20) 데이터 로딩
# ======================================
if "sigmat" in wb.sheetnames:
    ws_std20 = wb["sigmat"]
    std20_date_infos = []

    for col in range(3, ws_std20.max_column + 1):
        raw = ws_std20.cell(row=1, column=col).value
        if raw is None:
            continue
        dt = _to_datetime(raw)
        if dt is None:
            continue
        label = dt.strftime("%Y.%m.%d.")
        std20_date_infos.append((col, raw, dt, label))

    std20_date_infos = sorted(
        std20_date_infos,
        key=lambda x: (x[2] is None, x[2] or datetime.min)
    )

    total_std20_days = len(std20_date_infos)
    if total_std20_days > 0:
        show_std20 = min(st.session_state.show_days_std20, total_std20_days)
        start_idx = total_std20_days - show_std20
        selected_std20_infos = std20_date_infos[start_idx:]
        selected_std20_labels = [lbl for _, _, _, lbl in selected_std20_infos]

        std20_range_msg = (
            f"📅 표준편차 표시 범위: **{selected_std20_labels[0]} ~ {selected_std20_labels[-1]}** "
            f"(최근 {show_std20}일 / 전체 {total_std20_days}일)"
        )

        stock_info_norm = {}
        for code, name in stock_info.items():
            code_str = str(code).strip()
            if code_str.isdigit():
                code_str = code_str.zfill(6)
            stock_info_norm[code_str] = name

        label_to_col_std20 = {}
        for col in range(3, ws_std20.max_column + 1):
            raw = ws_std20.cell(row=1, column=col).value
            if raw is None:
                continue
            lbl = format_excel_date(raw)
            label_to_col_std20[lbl] = col

        std20_dict = {
            code: {"종목코드": code, "종목명": name}
            for code, name in stock_info_norm.items()
        }

        for r in range(2, ws_std20.max_row + 1):
            code = ws_std20.cell(row=r, column=2).value
            if code is None:
                continue
            code_str = str(code).strip()
            if code_str.isdigit():
                code_str = code_str.zfill(6)
            if code_str not in std20_dict:
                continue

            for lbl in selected_std20_labels:
                col_idx = label_to_col_std20.get(lbl)
                val = ws_std20.cell(row=r, column=col_idx).value if col_idx else None
                std20_dict[code_str][lbl] = val

        std20_df = pd.DataFrame.from_dict(std20_dict, orient="index").reset_index(drop=True)
        del std20_dict
        for lbl in selected_std20_labels:
            if lbl in std20_df.columns:
                std20_df[lbl] = pd.to_numeric(std20_df[lbl], errors="coerce")

        if is_index_target and "isigmat" in wb.sheetnames:
            ws_istd20 = wb["isigmat"]
            label_to_col_istd20 = {}
            for col in range(3, ws_istd20.max_column + 1):
                raw = ws_istd20.cell(row=1, column=col).value
                if raw is None:
                    continue
                lbl = format_excel_date(raw)
                label_to_col_istd20[lbl] = col

            rows_std20 = []
            for r in range(2, ws_istd20.max_row + 1):
                name = ws_istd20.cell(row=r, column=1).value
                code = ws_istd20.cell(row=r, column=2).value
                if not name or not code:
                    continue

                code_str = str(code).strip()
                if code_str.isdigit():
                    code_str = code_str.zfill(6)

                row_dict = {"업종명": str(name), "업종코드": code_str}
                for lbl in selected_std20_labels:
                    col_idx = label_to_col_istd20.get(lbl)
                    row_dict[lbl] = ws_istd20.cell(row=r, column=col_idx).value if col_idx else None
                rows_std20.append(row_dict)

            index_std20_rows = rows_std20

# ======================================
# 12.5. 매수량(순매수) 데이터 로딩
# ======================================
if is_index_target:
    netbuy_base_sheet = None
    for _, sheet_name in NETBUY_SHEET_MAP.items():
        if sheet_name in wb.sheetnames:
            netbuy_base_sheet = wb[sheet_name]
            break

    if netbuy_base_sheet is not None:
        netbuy_date_infos = []
        for col in range(3, netbuy_base_sheet.max_column + 1):
            raw = netbuy_base_sheet.cell(row=1, column=col).value
            if raw is None:
                continue
            dt = _to_datetime(raw)
            if dt is None:
                continue
            label = dt.strftime("%Y.%m.%d.")
            netbuy_date_infos.append((col, raw, dt, label))

        netbuy_date_infos = sorted(
            netbuy_date_infos,
            key=lambda x: (x[2] is None, x[2] or datetime.min)
        )

        total_netbuy_days = len(netbuy_date_infos)
        if total_netbuy_days > 0:
            show_buy = min(st.session_state.show_days_buy, total_netbuy_days)
            start_idx = total_netbuy_days - show_buy
            selected_buy_infos = netbuy_date_infos[start_idx:]
            selected_buy_labels = [lbl for _, _, _, lbl in selected_buy_infos]

            netbuy_range_msg = (
                f"📅 매수량 표시 범위: **{selected_buy_labels[0]} ~ {selected_buy_labels[-1]}** "
                f"(최근 {show_buy}일 / 전체 {total_netbuy_days}일)"
            )

            stock_info_norm = {}
            for code, name in stock_info.items():
                code_str = str(code).strip()
                if code_str.isdigit():
                    code_str = code_str.zfill(6)
                stock_info_norm[code_str] = name

            for investor_name, sheet_name in NETBUY_SHEET_MAP.items():
                if sheet_name not in wb.sheetnames:
                    continue
                ws_buy = wb[sheet_name]

                label_to_col = {}
                for col in range(3, ws_buy.max_column + 1):
                    raw = ws_buy.cell(row=1, column=col).value
                    if raw is None:
                        continue
                    lbl = format_excel_date(raw)
                    label_to_col[lbl] = col

                buy_dict = {
                    code: {"종목코드": code, "종목명": name}
                    for code, name in stock_info_norm.items()
                }

                for r in range(2, ws_buy.max_row + 1):
                    code = ws_buy.cell(row=r, column=2).value
                    if code is None:
                        continue
                    code_str = str(code).strip()
                    if code_str.isdigit():
                        code_str = code_str.zfill(6)
                    if code_str not in buy_dict:
                        continue

                    for lbl in selected_buy_labels:
                        col_idx = label_to_col.get(lbl)
                        val = ws_buy.cell(row=r, column=col_idx).value if col_idx else None
                        buy_dict[code_str][lbl] = val

                df_buy = pd.DataFrame.from_dict(buy_dict, orient="index").reset_index(drop=True)
                del buy_dict
                for lbl in selected_buy_labels:
                    if lbl in df_buy.columns:
                        df_buy[lbl] = pd.to_numeric(df_buy[lbl], errors="coerce")
                netbuy_df_map[investor_name] = df_buy

# ======================================
# 13. 엑셀 파일 닫기
# ======================================
wb.close()
# openpyxl 워크북과 워크시트 참조를 렌더링 전에 해제한다. 큰 파일에서 다른
# 카테고리로 전환할 때 이전/현재 실행의 메모리가 겹쳐 앱이 종료되는 것을 방지한다.
wb = None
base_ws = None
ws = None
ws_idx_raw = None
ws_idx = None
ws_std20 = None
ws_istd20 = None
netbuy_base_sheet = None
ws_buy = None
gc.collect()

# ======================================
# 14. 화면 구성 및 렌더링
# ======================================
view_options = ["1️⃣ 종합", "2️⃣ 지표별", "3️⃣ 원자료", "4️⃣ 매수량", "5️⃣ 표준편차"]
selected_view = st.radio(
    "조회 화면",
    view_options,
    key="selected_view",
    horizontal=True,
    label_visibility="collapsed",
)

# st.tabs는 선택하지 않은 탭의 대형 표까지 모두 렌더링한다. 현재 화면 하나만
# 렌더링해 카테고리 전환 시 이전/현재 표의 메모리가 겹치는 것을 방지한다.
if selected_view == "1️⃣ 종합":
    if indicator_df is None:
        st.warning("⚠️ 종합 데이터를 불러올 수 없습니다.")
    else:
        render_total_view(
            indicator_df,
            selected_labels,
            indicator_range_msg,
            total_days,
            index_df=index_df,
            index_metric_rows=index_metric_rows,
        )

elif selected_view == "2️⃣ 지표별":
    if indicator_df is None:
        st.warning("⚠️ 지표별 데이터를 불러올 수 없습니다.")
    else:
        render_metric_view(indicator_df, selected_labels, index_metric_map=index_metric_map)

elif selected_view == "3️⃣ 원자료":
    if close_df is None:
        st.warning("⚠️ 원자료(종가) 데이터를 불러올 수 없습니다.")
    else:
        render_raw_view(close_df, close_range_msg, total_close_days, index_close_rows=index_close_rows if is_index_target else None)

elif selected_view == "4️⃣ 매수량":
    if not is_index_target:
        st.info("KR_Stocks_Individual에서만 매수량 탭을 지원합니다.")
    elif not netbuy_df_map:
        st.warning("⚠️ 매수량(순매수) 시트가 없습니다. stock_history.py 실행 후 다시 확인해 주세요.")
    else:
        render_netbuy_view(netbuy_df_map, netbuy_range_msg, total_netbuy_days)

elif selected_view == "5️⃣ 표준편차":
    if std20_df is None:
        st.warning("⚠️ 표준편차(STD20) 시트가 없습니다. run_all_scores.py 또는 패치 워크플로 실행 후 다시 확인해 주세요.")
    else:
        render_std20_view(std20_df, std20_range_msg, total_std20_days, index_std20_rows=index_std20_rows if is_index_target else None)

st.markdown("---")
st.caption("Created by Alicia")
